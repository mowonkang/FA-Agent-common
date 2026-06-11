"""
FA 설비 유연율 표준화 엑셀 빌더 v2 (연도별 매트릭스 + raw 시트 추가)

- 입력: references/FA설비유연율 Data.md (393줄, 4 시트 raw)
- 출력: outputs/FA설비유연율_표준화_2026-05-21.xlsx (7 시트)
- 시트 구성
  · 01_대시보드
  · 02_MI_Lansing (표준화 + 연도별 매트릭스)
  · 03_MI_Lansing_raw (S1 raw 표 전체)
  · 04_GM2_각형 (표준화 + 연도별 매트릭스)
  · 05_GM2_raw
  · 06_HG (단일 합계)
  · 07_HG_raw
- 핵심
  · 라인 간 동일 컬럼·동일 산식·동일 단위 표준화
  · 연도별 매트릭스 = 세로(공정 + 설비) × 가로(연도) — 기술과제 누적 적용 효과
  · raw 시트 = S1 markdown 표 그대로 셀로 변환 (디테일 관리)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
import re

SRC = Path("/home/user/claude_FA-Agent_1/references/FA설비유연율 Data.md")
OUT = Path("/home/user/claude_FA-Agent_1/outputs/FA설비유연율_표준화_2026-05-21.xlsx")

# ============================================================
# 스타일
# ============================================================
# ── 모던 대시보드 팔레트 (진남 + 강조 파랑/초록/빨강) ──
COL_INK = "0F172A"       # 잉크 진남색 (메인 텍스트 / Hero 띠)
COL_INK_SOFT = "334155"  # 부드러운 진남회색 (헤더 진하기)
COL_GRAY_700 = "475569"
COL_GRAY_500 = "94A3B8"  # 보조 텍스트
COL_GRAY_300 = "CBD5E1"
COL_GRAY_100 = "F1F5F9"  # 카드 배경
COL_GRAY_50 = "F8FAFC"   # 매우 옅은 배경
COL_WHITE = "FFFFFF"
COL_ACCENT = "1E40AF"    # 강조 진파랑 (KPI 숫자)
COL_ACCENT_BG = "EFF6FF" # 강조 배경
COL_SUCCESS = "059669"   # 진초록 (변화율 ↑)
COL_SUCCESS_BG = "D1FAE5"
COL_DANGER = "DC2626"    # 빨강 (변화율 ↓)
COL_DANGER_BG = "FEE2E2"
COL_WARN = "D97706"      # 주황 (추정·주의)
COL_WARN_BG = "FEF3C7"

# 폰트 위계 (6단)
FONT_TITLE_W = Font(name="맑은 고딕", size=15, bold=True, color=COL_WHITE)
FONT_HERO_BIG = Font(name="맑은 고딕", size=28, bold=True, color=COL_ACCENT)   # KPI 숫자
FONT_HERO_MED = Font(name="맑은 고딕", size=14, bold=True, color=COL_INK)      # KPI 보조
FONT_HERO_LBL = Font(name="맑은 고딕", size=9, color=COL_GRAY_500)              # KPI 라벨
FONT_DELTA_POS = Font(name="맑은 고딕", size=12, bold=True, color=COL_SUCCESS)
FONT_DELTA_NEG = Font(name="맑은 고딕", size=12, bold=True, color=COL_DANGER)
FONT_H1 = Font(name="맑은 고딕", size=12, bold=True, color=COL_INK)
FONT_H2 = Font(name="맑은 고딕", size=10, bold=True, color=COL_INK)
FONT_H2_W = Font(name="맑은 고딕", size=10, bold=True, color=COL_WHITE)
FONT_BODY = Font(name="맑은 고딕", size=10, color=COL_INK)
FONT_RAW = Font(name="맑은 고딕", size=9, color=COL_INK)
FONT_RAW_HEAD = Font(name="맑은 고딕", size=9, bold=True, color=COL_INK)
FONT_INPUT = Font(name="맑은 고딕", size=10, color=COL_ACCENT, bold=True)
FONT_CALC = Font(name="맑은 고딕", size=10, color=COL_INK, bold=True)
FONT_KPI = Font(name="맑은 고딕", size=11, color=COL_ACCENT, bold=True)
FONT_LINK = Font(name="맑은 고딕", size=9, color=COL_GRAY_500)
FONT_NOTE = Font(name="맑은 고딕", size=9, italic=True, color=COL_GRAY_500)
FONT_FOOT = Font(name="맑은 고딕", size=8, bold=True, color=COL_GRAY_500)
FONT_FORMULA = Font(name="Consolas", size=9, italic=True, color=COL_GRAY_500)
FONT_TOTAL = Font(name="맑은 고딕", size=10, bold=True, color=COL_WHITE)
FONT_HERO = Font(name="맑은 고딕", size=11, bold=True, color=COL_INK)
FONT_TAG_WARN = Font(name="맑은 고딕", size=9, bold=True, color=COL_WARN)

# ── Fill (모던 대시보드) ──
FILL_TITLE = PatternFill("solid", fgColor=COL_INK)         # Hero 띠 진남 (흰글씨)
FILL_HERO = PatternFill("solid", fgColor=COL_ACCENT_BG)    # Hero Line 옅은 파랑
FILL_HEADER = PatternFill("solid", fgColor=COL_INK_SOFT)   # 헤더 진남 (흰글씨)
FILL_SUBHEAD = PatternFill("solid", fgColor=COL_GRAY_100)  # 서브헤더 옅은 회색
FILL_TOTAL = PatternFill("solid", fgColor=COL_INK)         # Total 진남 (흰글씨)
FILL_LGES = PatternFill("solid", fgColor=COL_GRAY_100)     # 라인 라벨 카드
FILL_S1 = PatternFill("solid", fgColor=COL_WHITE)          # 본문 화이트
FILL_S2 = PatternFill("solid", fgColor=COL_GRAY_50)        # 본문 alt
FILL_KEY = PatternFill("solid", fgColor=COL_ACCENT_BG)     # 강조 옅은 파랑
FILL_F1 = PatternFill("solid", fgColor=COL_ACCENT_BG)      # F1 사실 옅은 파랑
FILL_CARD = PatternFill("solid", fgColor=COL_GRAY_50)      # KPI 카드 배경
FILL_CARD_ACCENT = PatternFill("solid", fgColor=COL_ACCENT_BG)
FILL_SUCCESS = PatternFill("solid", fgColor=COL_SUCCESS_BG)
FILL_DANGER = PatternFill("solid", fgColor=COL_DANGER_BG)
FILL_WARN = PatternFill("solid", fgColor=COL_WARN_BG)
# 연도별 단일 톤 농도 (옅은 파랑→진한 파랑)
FILL_Y25 = PatternFill("solid", fgColor="F8FAFC")
FILL_Y26 = PatternFill("solid", fgColor="E0E7FF")
FILL_Y27 = PatternFill("solid", fgColor="C7D2FE")
FILL_Y28 = PatternFill("solid", fgColor="A5B4FC")
# 연도별 단일 톤 농도 (파랑 옅음→진함)
FILL_Y25 = PatternFill("solid", fgColor="F0F4F8")
FILL_Y26 = PatternFill("solid", fgColor="DCE7F1")
FILL_Y27 = PatternFill("solid", fgColor="C7DAEA")
FILL_Y28 = PatternFill("solid", fgColor="B3CDE3")
FILL_RAW = PatternFill("solid", fgColor="FAFAFA")
FILL_RAW_HEAD = PatternFill("solid", fgColor="E5E7EB")

THIN = Side(style="thin", color="808080")
HAIR = Side(style="hair", color="C0C0C0")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
RAW_BOX = Border(left=HAIR, right=HAIR, top=HAIR, bottom=HAIR)
ALIGN_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_L = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_R = Alignment(horizontal="right", vertical="center", wrap_text=True)
ALIGN_RAW = Alignment(horizontal="left", vertical="top", wrap_text=True)


def set_cell(ws, row, col, value, font=None, fill=None, align=None, border=BOX, fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    if font: c.font = font
    if fill: c.fill = fill
    if align: c.alignment = align
    if border: c.border = border
    if fmt: c.number_format = fmt
    return c


def title_band(ws, row, text, span):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = FONT_TITLE_W
    c.fill = FILL_TITLE
    c.alignment = ALIGN_C
    ws.row_dimensions[row].height = 26


# ============================================================
# 데이터: 표준화 (S1 서머리)
# ============================================================
# (공정, '25 Total, '25 유연, '25 고정, '25 비고, '28 Total, '28 유연, '28 고정, '28 비고)
MI_LANSING = [
    ("전극",   275.384725, 98.143,            177.241725,
     "점보롤 스토커 고정 설비, 브릿지 물류",
     258.853725, 258.853725, 0.0,
     "점보롤 스토커 고정 설비, 브릿지 물류 → SMC, 밀폐형 AMR"),
    ("조립",   610.89167125, 417.34,           193.55167125,
     "팬케익 Rack → SMC",
     596.93167125, 448.33, 148.60167125,
     "잔여: PKG 배출 CNV, 조립 원부자재 OHMS"),
    ("활성화", 443.11, 99.80394146,             343.30605854,
     "활성화 Rack → SMC, MMF",
     423.87, 272.96394146, 150.90605854,
     "잔여: Aging Lift, 연결 물류"),
    ("출하",   132.60793375, 72.974,            59.63903125,
     "출하 Rack → SMC, MMF",
     132.62, 91.39, 41.23,
     "잔여: 포장기, 세척기"),
]

GM2 = [
    ("전극",   94.4168, 87.2888, 7.128,
     "모바일셔틀 Rack",
     93.704, 93.704, 0.0,
     "모바일셔틀 Rack → SMC"),
    ("조립",   678.4261264, 341.53904, 336.8870864,
     "팬케익, 자재 창고 Rack → SMC",
     658.42825076, 544.75654076, 113.67171,
     "잔여: OHMS Rail, Lift, 포장해체기"),
    ("활성화", 447.6610545, 188.42176, 259.2392945,
     "활성화 Rack → SMC, MMF",
     436.12839705, 317.27477705, 118.8536200,
     "잔여: Encloser, HVC/충방전기 Stocker, 연결 물류"),
    ("출하",   105.245446, 42.60184, 62.643606,
     "출하 Rack → SMC, MMF",
     104.478406, 49.5052, 54.973206,
     "잔여: 포장기, 세척기, 연결물류"),
]

HG_TOTAL = {
    "예산 총액": 3465.84,
    "설비비": 1968.84998,
    "설치비": 1146.19002,
    "유연설비": 834.44,
    "고정설비": 2631.40,
    "유연율": 0.2408,
    "출처": "S1 [322~393행]",
    "비고": "raw 시트가 설비 단위(AMR/모바일셔틀/Aging/OHT 등) 로 분해. 4공정 매트릭스는 본 자료에 미공시 — 합계만 박제",
}

# HG 라인의 설비 단위 → 4공정 추정 분해 (S1 [322~393행])
# AMR/모바일셔틀 = 유연 / Aging+CNV/OHT/포장설비/포일이재기/트레이세척기 = 고정
# 행 8(AMR 전극 175.4) + 행 9(모바일셔틀 전극 731.9) + 행 16(포일이재기 23.64)
# 행 17(AMR 조립 270.9) + 행 18(트레이세척기 8.2)
# 행 19(AMR 활성화/출하 115.4) + 행 10(모바일셔틀 활성화/출하 197) + 행 20(포장설비 75.3) + 행 21(Aging+CNV 1034.5) + 행 22(OHT 510.9) + 행 23(75.2) + 행 24(247.5)
HG_BY_PROCESS = [
    # (공정, Total, 유연, 고정, 비고)
    ("전극",   175.4 + 731.9 + 23.64, 175.4 + 731.9, 23.64,
     "AMR 175.4 + 모바일셔틀 731.9 (유연) / 포일이재기 23.64 (고정)"),
    ("조립",   270.9 + 8.2, 270.9, 8.2,
     "AMR 270.9 (유연) / 트레이세척기 8.2 (고정)"),
    ("활성화", 1034.5 + 510.9, 0, 1034.5 + 510.9,
     "Aging+CNV 1034.5 + OHT 510.9 (모두 고정 분류, 설치비 큼)"),
    ("출하",   115.4 + 197 + 75.3 + 75.2 + 247.5, 115.4 + 197, 75.3 + 75.2 + 247.5,
     "AMR 115.4 + 모바일셔틀 197 (유연) / 포장설비 75.3 + OHT 이관 75.2 + 추가 자동화 247.5 (고정)"),
]

# HG '25~'28 추정 시계열 (추정 — 사내 사실 부재)
# 가정: '24 24% 기준 → MI Lansing 패턴 적용 (활성화 SMC·MMF 미적용 가정으로 더 낮은 곡선)
# '26 전극 SMC 적용 시 전극 100%, 활성화는 미개선 → Total ~35%
# '27 MMF 적용 시 활성화 60% → Total ~50%
# '28 밀폐형 AMR 시 출하 70% → Total ~58%
F1_TIMESERIES = {
    "ESMI Lansing": (0.47, 0.61, 0.75, 0.76),
    "ESGM2 각형":   (0.50, 0.62, 0.74, 0.78),
    "ESHG":         (0.24, None, None, None),  # 24년 실측만 사내 사실
}
# HG 별도 추정 (참고용, 가정 명시)
HG_ESTIMATE = {
    "'25": 0.27,  # 24→25 점진 (혁신물류2.0 일부 반영)
    "'26": 0.42,  # 전극 SMC 가정 (활성화 미개선)
    "'27": 0.55,  # MMF 가정 (활성화 60%)
    "'28": 0.62,  # 밀폐형 AMR 가정 (출하 70%)
}
# 연도별 주요 과제 — S1 [44행]
F1_TASKS = {
    "'25": "혁신물류2.0",
    "'26": "전극 SMC",
    "'27": "MMF (활성화 SMC)",
    "'28": "밀폐형 AMR",
}

# 연도별 공정 유연율 모델 (가정) — 라인 시트의 §연도별 매트릭스용
# 각 공정에 어떤 연도의 과제가 적용되는지 + 그 효과
# '25 = S1 시나리오 1, '28 = S1 시나리오 2, 중간 '26·'27 은 선형 누적 가정
# 형식: 라인 → {공정: [('25 유연율, '26, '27, '28)], "_task": {공정: 적용 연도}}
MI_LANSING_YEARLY = {
    "전극":   (0.356, 1.000, 1.000, 1.000),   # '26 전극 SMC 적용
    "조립":   (0.683, 0.700, 0.730, 0.751),   # 점진적 (SMC 외 잔여)
    "활성화": (0.225, 0.300, 0.644, 0.644),   # '27 MMF 적용
    "출하":   (0.550, 0.560, 0.600, 0.689),   # '28 밀폐형 AMR 적용
}
GM2_YEARLY = {
    "전극":   (0.925, 1.000, 1.000, 1.000),   # '26 모바일셔틀 → SMC
    "조립":   (0.503, 0.680, 0.750, 0.827),   # '26 단계 SMC 확대
    "활성화": (0.421, 0.500, 0.727, 0.727),
    "출하":   (0.405, 0.420, 0.450, 0.474),
}
# 공정별 적용 연도 매핑 (비고)
TASK_BY_PROCESS = {
    "전극":   "'26 전극 SMC (모바일셔틀/점보롤 Rack → SMC 전환)",
    "조립":   "'25~'28 점진 (팬케익 Rack → SMC, PKG 배출 CNV/OHMS 잔여)",
    "활성화": "'27 MMF (활성화 Rack → SMC, MMF 적용)",
    "출하":   "'28 밀폐형 AMR (출하 Rack → SMC, 밀폐형 AMR)",
}


# ============================================================
# Markdown 표 파서 — S1 raw 시트를 워크시트로 옮김
# ============================================================
def parse_md_section(md_text, sheet_name):
    """### Sheet: <sheet_name> 으로 시작하는 표 영역을 파싱하여 2D 리스트 반환."""
    lines = md_text.splitlines()
    start = None
    for i, line in enumerate(lines):
        if line.strip().startswith("### Sheet:") and sheet_name in line:
            start = i + 1
            break
    if start is None:
        return []
    end = len(lines)
    for j in range(start, len(lines)):
        if lines[j].strip().startswith("### Sheet:"):
            end = j
            break
    rows = []
    for line in lines[start:end]:
        s = line.strip()
        if not s.startswith("|"):
            continue
        # 구분선(---) 행 스킵
        if re.fullmatch(r"\|[\s\-:|]+\|", s):
            continue
        cells = [c.strip() for c in s.split("|")[1:-1]]
        rows.append(cells)
    return rows


def write_raw_sheet(ws, title, rows, source_line_range):
    """parse_md_section 결과를 워크시트에 그대로 옮김."""
    title_band(ws, 1, title, max(8, min(len(rows[0]) if rows else 8, 30)))
    set_cell(ws, 3, 1, "출처", font=FONT_H2, fill=FILL_HEADER, align=ALIGN_L)
    span = min(len(rows[0]) if rows else 8, 30)
    ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=span)
    set_cell(ws, 3, 2, f"references/FA설비유연율 Data.md {source_line_range} — S1 raw 표 그대로 옮김 (전체 컬럼 보존)",
             font=FONT_LINK, align=ALIGN_L)
    set_cell(ws, 4, 1, "사용 안내", font=FONT_H2, fill=FILL_HEADER, align=ALIGN_L)
    ws.merge_cells(start_row=4, start_column=2, end_row=4, end_column=span)
    set_cell(ws, 4, 2, "본 시트는 raw 데이터 보존용. 디테일 관리 목적. 표준화 분석은 02·04·06 시트 참조.",
             font=FONT_NOTE, align=ALIGN_L)

    if not rows:
        set_cell(ws, 6, 1, "(raw 데이터 부재)", font=FONT_NOTE, align=ALIGN_L)
        return

    start_row = 6
    for ri, row in enumerate(rows):
        for ci, val in enumerate(row):
            # 숫자 변환 시도
            v = val
            try:
                if val and val != "nan":
                    v = float(val)
                    if v.is_integer() and abs(v) < 1e10:
                        v = int(v)
                else:
                    v = ""
            except (ValueError, TypeError):
                v = val if val != "nan" else ""

            is_header = (ri == 0)
            font = FONT_RAW_HEAD if is_header else FONT_RAW
            fill = FILL_RAW_HEAD if is_header else FILL_RAW
            align = ALIGN_C if is_header else ALIGN_RAW
            set_cell(ws, start_row + ri, ci + 1, v, font=font, fill=fill,
                     align=align, border=RAW_BOX,
                     fmt="#,##0.0000" if isinstance(v, float) else None)

    # 컬럼 폭 자동 조정 (최대 30개 컬럼)
    ncols = min(len(rows[0]), 30) if rows else 0
    for ci in range(1, ncols + 1):
        col = get_column_letter(ci)
        # 컬럼별 최대 길이
        max_len = 0
        for ri in range(min(len(rows), 100)):
            try:
                v = str(rows[ri][ci - 1])
                if len(v) > max_len:
                    max_len = len(v)
            except IndexError:
                pass
        ws.column_dimensions[col].width = min(max(8, max_len * 0.9), 25)


# ============================================================
# 워크북 빌드
# ============================================================
wb = Workbook()
wb.remove(wb.active)
md_text = SRC.read_text(encoding="utf-8")

# ============================================================
# Sheet 1 — 대시보드
# ============================================================
ws = wb.create_sheet("01_대시보드")
# 12 컬럼 (A~L) — 3 카드 × 4 컬럼 정렬
widths = [13, 11, 13, 11, 13, 11, 13, 11, 13, 11, 13, 13]
for i, w in enumerate(widths):
    ws.column_dimensions[get_column_letter(i + 1)].width = w

YEAR_FILLS = [FILL_Y25, FILL_Y26, FILL_Y27, FILL_Y28]
YEAR_LABELS = ["'25년", "'26년", "'27년", "'28년"]
TASKS_BY_YEAR = ["혁신물류2.0", "전극 SMC", "MMF (활성화 SMC)", "밀폐형 AMR"]

# ───────────────────────────────────────────
# 행 1 — Hero 타이틀 (진남 띠)
# ───────────────────────────────────────────
ws.merge_cells("A1:L1")
c1 = ws.cell(row=1, column=1, value="  FA 설비 유연율 — 라인 통합 대시보드")
c1.font = FONT_TITLE_W
c1.fill = FILL_TITLE
c1.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 38

# 행 2 — KPI Hero Line (한 줄 요약, 옅은 파랑)
ws.merge_cells("A2:L2")
set_cell(ws, 2, 1,
         "  MI Lansing  47% → 76% (+29%p)    │    GM2 각형  50% → 78% (+28%p)    │    HG  24% → 62% (+38%p, 추정)    │    4년 누적: 혁신물류2.0 → 전극 SMC → MMF → 밀폐형 AMR",
         font=FONT_HERO, fill=FILL_HERO, align=Alignment(horizontal="left", vertical="center"))
ws.row_dimensions[2].height = 30

# 행 3 — 메타 1줄
ws.merge_cells("A3:L3")
set_cell(ws, 3, 1,
         "  작성: 2026-05-21  ·  정의: F1 사내 사실  ·  산식: 유연율 = Σ(유연설비)/Σ(전체)  ·  유연 기준: 셋업비 비중 30% 이하  ·  출처: S1 references/FA설비유연율 Data.md",
         font=FONT_NOTE, fill=FILL_S2, align=Alignment(horizontal="left", vertical="center"))
ws.row_dimensions[3].height = 22

# ───────────────────────────────────────────
# §0. KPI Hero Cards (행 5~9, 3장 가로 배치)
# ───────────────────────────────────────────
# 각 카드 4 컬럼 (A-D / E-H / I-L)
cards = [
    {
        "label": "MI Lansing",
        "kpi": "='02_MI_Lansing'!F29",
        "delta": "='02_MI_Lansing'!F29-'02_MI_Lansing'!C29",
        "start": "='02_MI_Lansing'!C29",
        "memo": "미시간 Lansing 각형 · 1차 양산",
        "tag": "사실",
        "tag_fill": FILL_F1,
        "tag_font": FONT_NOTE,
    },
    {
        "label": "GM2 각형",
        "kpi": "='04_GM2_각형'!F29",
        "delta": "='04_GM2_각형'!F29-'04_GM2_각형'!C29",
        "start": "='04_GM2_각형'!C29",
        "memo": "GM2 신규 · 양산 대기",
        "tag": "사실",
        "tag_fill": FILL_F1,
        "tag_font": FONT_NOTE,
    },
    {
        "label": "HG",
        "kpi": HG_ESTIMATE["'28"],
        "delta": HG_ESTIMATE["'28"] - 0.24,
        "start": 0.24,
        "memo": "HG 라인 · 활성화 Aging 잔존",
        "tag": "추정",
        "tag_fill": FILL_WARN,
        "tag_font": FONT_TAG_WARN,
    },
]

for ci, card in enumerate(cards):
    c0 = 1 + ci * 4   # A=1 / E=5 / I=9
    # 카드 배경 박스 — 행 5~9
    for r in range(5, 10):
        for col in range(c0, c0 + 4):
            ws.cell(row=r, column=col).fill = FILL_CARD
    # 행 5: 라벨 + 태그
    ws.merge_cells(start_row=5, start_column=c0, end_row=5, end_column=c0+2)
    set_cell(ws, 5, c0, "  " + card["label"], font=FONT_H1, fill=FILL_CARD, align=Alignment(horizontal="left", vertical="center"), border=None)
    set_cell(ws, 5, c0+3, card["tag"], font=card["tag_font"], fill=card["tag_fill"], align=ALIGN_C, border=None)
    ws.row_dimensions[5].height = 24

    # 행 6-7: 큰 KPI 숫자 + 변화율
    ws.merge_cells(start_row=6, start_column=c0, end_row=7, end_column=c0+1)
    set_cell(ws, 6, c0, card["kpi"], font=FONT_HERO_BIG, fill=FILL_CARD,
             align=Alignment(horizontal="center", vertical="center"), border=None, fmt="0%")
    ws.row_dimensions[6].height = 22
    ws.row_dimensions[7].height = 22
    # 우측: '28 목표 라벨 + 변화율
    ws.merge_cells(start_row=6, start_column=c0+2, end_row=6, end_column=c0+3)
    set_cell(ws, 6, c0+2, "'28 목표", font=FONT_HERO_LBL, fill=FILL_CARD,
             align=Alignment(horizontal="left", vertical="bottom"), border=None)
    ws.merge_cells(start_row=7, start_column=c0+2, end_row=7, end_column=c0+3)
    set_cell(ws, 7, c0+2, card["delta"], font=FONT_DELTA_POS, fill=FILL_CARD,
             align=Alignment(horizontal="left", vertical="top"), border=None, fmt="+0.0%\" vs '25\"")

    # 행 8: "from 47% ('25)"
    ws.merge_cells(start_row=8, start_column=c0, end_row=8, end_column=c0+3)
    set_cell(ws, 8, c0, card["start"], font=FONT_HERO_LBL, fill=FILL_CARD,
             align=Alignment(horizontal="center", vertical="center"), border=None,
             fmt="\"  시작 \"0%\" ('25년 기준)\"")
    # 행 9: 메모
    ws.merge_cells(start_row=9, start_column=c0, end_row=9, end_column=c0+3)
    set_cell(ws, 9, c0, "  " + card["memo"], font=FONT_NOTE, fill=FILL_CARD,
             align=Alignment(horizontal="left", vertical="center"), border=None)

# 카드 간 시각적 구분 (얇은 보더)
for ci in range(len(cards)):
    c0 = 1 + ci * 4
    for r in range(5, 10):
        left_cell = ws.cell(row=r, column=c0)
        right_cell = ws.cell(row=r, column=c0+3)
        left_cell.border = Border(left=Side(style="medium", color=COL_ACCENT))

# ───────────────────────────────────────────
# §1. 라인 × 연도 매트릭스 (행 11~17)
# ───────────────────────────────────────────
r = 11
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
set_cell(ws, r, 1, "§1. 라인 × 연도 유연율 매트릭스", font=FONT_H1, fill=FILL_SUBHEAD, align=ALIGN_L)
ws.row_dimensions[r].height = 22

# 행 12: 헤더 (연도)
r = 12
ws.merge_cells(start_row=r, start_column=1, end_row=r+1, end_column=1)
set_cell(ws, r, 1, "라인", font=FONT_H2_W, fill=FILL_HEADER, align=ALIGN_C)
for j, y in enumerate(YEAR_LABELS):
    set_cell(ws, r, j + 2, y, font=FONT_H2_W, fill=FILL_HEADER, align=ALIGN_C)
ws.merge_cells(start_row=r, start_column=6, end_row=r+1, end_column=6)
set_cell(ws, r, 6, "개선폭", font=FONT_H2_W, fill=FILL_HEADER, align=ALIGN_C)
ws.merge_cells(start_row=r, start_column=7, end_row=r, end_column=8)
set_cell(ws, r, 7, "투자비 '25 (억원)", font=FONT_H2_W, fill=FILL_HEADER, align=ALIGN_C)
ws.merge_cells(start_row=r, start_column=9, end_row=r, end_column=10)
set_cell(ws, r, 9, "투자비 '28 (억원)", font=FONT_H2_W, fill=FILL_HEADER, align=ALIGN_C)
ws.merge_cells(start_row=r, start_column=11, end_row=r+1, end_column=12)
set_cell(ws, r, 11, "라인 특성", font=FONT_H2_W, fill=FILL_HEADER, align=ALIGN_C)

# 행 13: 헤더 2
r = 13
for j, t in enumerate(TASKS_BY_YEAR):
    set_cell(ws, r, j + 2, t, font=FONT_NOTE, fill=YEAR_FILLS[j], align=ALIGN_C)
set_cell(ws, r, 7, "Total", font=FONT_NOTE, fill=FILL_S2, align=ALIGN_C)
set_cell(ws, r, 8, "유연", font=FONT_NOTE, fill=FILL_S2, align=ALIGN_C)
set_cell(ws, r, 9, "Total", font=FONT_NOTE, fill=FILL_S2, align=ALIGN_C)
set_cell(ws, r, 10, "유연", font=FONT_NOTE, fill=FILL_S2, align=ALIGN_C)
ws.row_dimensions[12].height = 26
ws.row_dimensions[13].height = 22

# 행 14: MI Lansing
r = 14
set_cell(ws, r, 1, "MI Lansing", font=FONT_H2, fill=FILL_LGES, align=ALIGN_L)
for j, col in enumerate(["C", "D", "E", "F"]):
    set_cell(ws, r, j + 2, f"='02_MI_Lansing'!{col}29",
             font=FONT_CALC, fill=YEAR_FILLS[j], align=ALIGN_C, fmt="0.0%")
set_cell(ws, r, 6, "=E14-B14", font=FONT_DELTA_POS, fill=FILL_SUCCESS, align=ALIGN_C, fmt="+0.0%;-0.0%")
set_cell(ws, r, 7, "='02_MI_Lansing'!C12", font=FONT_CALC, fill=FILL_S2, align=ALIGN_R, fmt="#,##0")
set_cell(ws, r, 8, "='02_MI_Lansing'!D12", font=FONT_CALC, fill=FILL_S2, align=ALIGN_R, fmt="#,##0")
set_cell(ws, r, 9, "='02_MI_Lansing'!C20", font=FONT_CALC, fill=FILL_S1, align=ALIGN_R, fmt="#,##0")
set_cell(ws, r, 10, "='02_MI_Lansing'!D20", font=FONT_CALC, fill=FILL_S1, align=ALIGN_R, fmt="#,##0")
ws.merge_cells(start_row=r, start_column=11, end_row=r, end_column=12)
set_cell(ws, r, 11, "미시간 Lansing · 1차 양산", font=FONT_NOTE, fill=FILL_S1, align=ALIGN_L)

# 행 15: GM2
r = 15
set_cell(ws, r, 1, "GM2 각형", font=FONT_H2, fill=FILL_LGES, align=ALIGN_L)
for j, col in enumerate(["C", "D", "E", "F"]):
    set_cell(ws, r, j + 2, f"='04_GM2_각형'!{col}29",
             font=FONT_CALC, fill=YEAR_FILLS[j], align=ALIGN_C, fmt="0.0%")
set_cell(ws, r, 6, "=E15-B15", font=FONT_DELTA_POS, fill=FILL_SUCCESS, align=ALIGN_C, fmt="+0.0%;-0.0%")
set_cell(ws, r, 7, "='04_GM2_각형'!C12", font=FONT_CALC, fill=FILL_S2, align=ALIGN_R, fmt="#,##0")
set_cell(ws, r, 8, "='04_GM2_각형'!D12", font=FONT_CALC, fill=FILL_S2, align=ALIGN_R, fmt="#,##0")
set_cell(ws, r, 9, "='04_GM2_각형'!C20", font=FONT_CALC, fill=FILL_S1, align=ALIGN_R, fmt="#,##0")
set_cell(ws, r, 10, "='04_GM2_각형'!D20", font=FONT_CALC, fill=FILL_S1, align=ALIGN_R, fmt="#,##0")
ws.merge_cells(start_row=r, start_column=11, end_row=r, end_column=12)
set_cell(ws, r, 11, "GM2 신규 · 양산 대기", font=FONT_NOTE, fill=FILL_S1, align=ALIGN_L)

# 행 16: HG (추정)
r = 16
set_cell(ws, r, 1, "HG (추정)", font=FONT_H2, fill=FILL_LGES, align=ALIGN_L)
hg_vals = [HG_ESTIMATE["'25"], HG_ESTIMATE["'26"], HG_ESTIMATE["'27"], HG_ESTIMATE["'28"]]
for j, v in enumerate(hg_vals):
    set_cell(ws, r, j + 2, v, font=FONT_INPUT, fill=YEAR_FILLS[j], align=ALIGN_C, fmt="0.0%")
set_cell(ws, r, 6, "=E16-B16", font=FONT_DELTA_POS, fill=FILL_SUCCESS, align=ALIGN_C, fmt="+0.0%;-0.0%")
set_cell(ws, r, 7, "='06_HG'!C8", font=FONT_CALC, fill=FILL_S2, align=ALIGN_R, fmt="#,##0")
set_cell(ws, r, 8, "='06_HG'!D8", font=FONT_CALC, fill=FILL_S2, align=ALIGN_R, fmt="#,##0")
set_cell(ws, r, 9, "='06_HG'!C8", font=FONT_CALC, fill=FILL_S1, align=ALIGN_R, fmt="#,##0")
set_cell(ws, r, 10, "='06_HG'!C8*E16", font=FONT_CALC, fill=FILL_S1, align=ALIGN_R, fmt="#,##0")
ws.merge_cells(start_row=r, start_column=11, end_row=r, end_column=12)
set_cell(ws, r, 11, "HG · 활성화 Aging 잔존 (추정)", font=FONT_TAG_WARN, fill=FILL_WARN, align=ALIGN_L)

# 행 17: F1 검증
r = 17
set_cell(ws, r, 1, "F1 사실 (검증)", font=FONT_H2, fill=FILL_F1, align=ALIGN_L)
for j, v in enumerate(F1_TIMESERIES["ESMI Lansing"]):
    set_cell(ws, r, j + 2, v, font=FONT_INPUT, fill=FILL_F1, align=ALIGN_C, fmt="0.0%")
set_cell(ws, r, 6, "=E17-B17", font=FONT_DELTA_POS, fill=FILL_F1, align=ALIGN_C, fmt="+0.0%;-0.0%")
ws.merge_cells(start_row=r, start_column=7, end_row=r, end_column=12)
set_cell(ws, r, 7, "ESMI Lansing F1 시계열 (사내 사실) — 행 14 수식 결과와 ±2%p 이내 정합 검증",
         font=FONT_LINK, fill=FILL_F1, align=ALIGN_L)

# ───────────────────────────────────────────
# §2. 연도별 변화 카드 (행 19~25)
# ───────────────────────────────────────────
r = 19
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
set_cell(ws, r, 1, "§2. 연도별 변화 카드 (★ 각 연도 핵심 과제 적용 → 라인별 유연율 변화)",
         font=FONT_H1, fill=FILL_SUBHEAD, align=ALIGN_L)
ws.row_dimensions[r].height = 22

# 행 20: 카드 헤더 (연도+과제)
r = 20
set_cell(ws, r, 1, "구분", font=FONT_H2_W, fill=FILL_HEADER, align=ALIGN_C)
for j, (y, t) in enumerate(zip(YEAR_LABELS, TASKS_BY_YEAR)):
    col_start = 2 + j * 2
    ws.merge_cells(start_row=r, start_column=col_start, end_row=r, end_column=col_start+1)
    set_cell(ws, r, col_start, f"{y}  {t}", font=FONT_H2_W, fill=FILL_HEADER, align=ALIGN_C)
ws.merge_cells(start_row=r, start_column=10, end_row=r, end_column=12)
set_cell(ws, r, 10, "'25 → '28 누적", font=FONT_H2_W, fill=FILL_HEADER, align=ALIGN_C)

# 행 21: MI Lansing
r = 21
set_cell(ws, r, 1, "MI Lansing", font=FONT_H2, fill=FILL_LGES, align=ALIGN_L)
for j in range(4):
    c_pair = [2 + j*2, 3 + j*2]
    src_col = ["C", "D", "E", "F"][j]
    set_cell(ws, r, c_pair[0], f"='02_MI_Lansing'!{src_col}29",
             font=FONT_CALC, fill=YEAR_FILLS[j], align=ALIGN_C, fmt="0.0%")
    if j == 0:
        set_cell(ws, r, c_pair[1], "기준", font=FONT_NOTE, fill=YEAR_FILLS[j], align=ALIGN_C)
    else:
        prev_col = ["C", "D", "E"][j-1]
        set_cell(ws, r, c_pair[1], f"='02_MI_Lansing'!{src_col}29-'02_MI_Lansing'!{prev_col}29",
                 font=FONT_DELTA_POS, fill=YEAR_FILLS[j], align=ALIGN_C, fmt="+0.0%;-0.0%")
ws.merge_cells(start_row=r, start_column=10, end_row=r, end_column=12)
set_cell(ws, r, 10, "=F14-B14", font=FONT_DELTA_POS, fill=FILL_SUCCESS, align=ALIGN_C, fmt="+0.0%;-0.0%")

# 행 22: GM2
r = 22
set_cell(ws, r, 1, "GM2 각형", font=FONT_H2, fill=FILL_LGES, align=ALIGN_L)
for j in range(4):
    c_pair = [2 + j*2, 3 + j*2]
    src_col = ["C", "D", "E", "F"][j]
    set_cell(ws, r, c_pair[0], f"='04_GM2_각형'!{src_col}29",
             font=FONT_CALC, fill=YEAR_FILLS[j], align=ALIGN_C, fmt="0.0%")
    if j == 0:
        set_cell(ws, r, c_pair[1], "기준", font=FONT_NOTE, fill=YEAR_FILLS[j], align=ALIGN_C)
    else:
        prev_col = ["C", "D", "E"][j-1]
        set_cell(ws, r, c_pair[1], f"='04_GM2_각형'!{src_col}29-'04_GM2_각형'!{prev_col}29",
                 font=FONT_DELTA_POS, fill=YEAR_FILLS[j], align=ALIGN_C, fmt="+0.0%;-0.0%")
ws.merge_cells(start_row=r, start_column=10, end_row=r, end_column=12)
set_cell(ws, r, 10, "=F15-B15", font=FONT_DELTA_POS, fill=FILL_SUCCESS, align=ALIGN_C, fmt="+0.0%;-0.0%")

# 행 23: HG (추정)
r = 23
set_cell(ws, r, 1, "HG (추정)", font=FONT_H2, fill=FILL_LGES, align=ALIGN_L)
for j in range(4):
    c_pair = [2 + j*2, 3 + j*2]
    cur_letter = ["B", "C", "D", "E"][j]
    set_cell(ws, r, c_pair[0], f"={cur_letter}16", font=FONT_CALC, fill=YEAR_FILLS[j], align=ALIGN_C, fmt="0.0%")
    if j == 0:
        set_cell(ws, r, c_pair[1], "기준", font=FONT_NOTE, fill=YEAR_FILLS[j], align=ALIGN_C)
    else:
        prev_letter = ["B", "C", "D"][j-1]
        set_cell(ws, r, c_pair[1], f"={cur_letter}16-{prev_letter}16",
                 font=FONT_DELTA_POS, fill=YEAR_FILLS[j], align=ALIGN_C, fmt="+0.0%;-0.0%")
ws.merge_cells(start_row=r, start_column=10, end_row=r, end_column=12)
set_cell(ws, r, 10, "=F16-B16", font=FONT_DELTA_POS, fill=FILL_SUCCESS, align=ALIGN_C, fmt="+0.0%;-0.0%")

# 행 24: 견인 공정 (라인 공통)
r = 24
set_cell(ws, r, 1, "견인 공정", font=FONT_H2, fill=FILL_SUBHEAD, align=ALIGN_L)
drivers = ["조립 부분 개선", "전극 100% 전환", "활성화 64~73%", "출하 47~69%"]
for j, d in enumerate(drivers):
    col_start = 2 + j*2
    ws.merge_cells(start_row=r, start_column=col_start, end_row=r, end_column=col_start+1)
    set_cell(ws, r, col_start, d, font=FONT_BODY, fill=YEAR_FILLS[j], align=ALIGN_C)
ws.merge_cells(start_row=r, start_column=10, end_row=r, end_column=12)
set_cell(ws, r, 10, "전극 → 활성화 → 출하 순차 전환", font=FONT_NOTE, fill=FILL_KEY, align=ALIGN_C)

# 행 25: 변경 설비
r = 25
set_cell(ws, r, 1, "변경 설비", font=FONT_H2, fill=FILL_SUBHEAD, align=ALIGN_L)
equipment = ["팬케익 Rack→SMC", "점보롤·모바일셔틀→SMC", "활성화 Rack→SMC, MMF", "출하 Rack→SMC, AMR"]
for j, e in enumerate(equipment):
    col_start = 2 + j*2
    ws.merge_cells(start_row=r, start_column=col_start, end_row=r, end_column=col_start+1)
    set_cell(ws, r, col_start, e, font=FONT_NOTE, fill=YEAR_FILLS[j], align=ALIGN_C)
ws.merge_cells(start_row=r, start_column=10, end_row=r, end_column=12)
set_cell(ws, r, 10, "고정 설비 → 유연 설비", font=FONT_NOTE, fill=FILL_KEY, align=ALIGN_C)
ws.row_dimensions[25].height = 28

# ───────────────────────────────────────────
# §3. 공정 × 연도 히트맵 (행 27~32, MI Lansing + GM2)
# ───────────────────────────────────────────
r = 27
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
set_cell(ws, r, 1, "§3. 공정 × 연도 히트맵 (좌: MI Lansing  /  우: GM2 각형)",
         font=FONT_H1, fill=FILL_SUBHEAD, align=ALIGN_L)
ws.row_dimensions[r].height = 22

# 행 28: 그룹 라벨
r = 28
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
set_cell(ws, r, 1, "MI Lansing", font=FONT_H2, fill=FILL_LGES, align=ALIGN_C)
ws.merge_cells(start_row=r, start_column=7, end_row=r, end_column=12)
set_cell(ws, r, 7, "GM2 각형", font=FONT_H2, fill=FILL_LGES, align=ALIGN_C)

# 행 29: 컬럼 헤더
r = 29
set_cell(ws, r, 1, "공정", font=FONT_H2_W, fill=FILL_HEADER, align=ALIGN_C)
for j, y in enumerate(YEAR_LABELS):
    set_cell(ws, r, j + 2, y, font=FONT_H2_W, fill=FILL_HEADER, align=ALIGN_C)
set_cell(ws, r, 6, "Δ%p", font=FONT_H2_W, fill=FILL_HEADER, align=ALIGN_C)
set_cell(ws, r, 7, "공정", font=FONT_H2_W, fill=FILL_HEADER, align=ALIGN_C)
for j, y in enumerate(YEAR_LABELS):
    set_cell(ws, r, j + 8, y, font=FONT_H2_W, fill=FILL_HEADER, align=ALIGN_C)
set_cell(ws, r, 12, "Δ%p", font=FONT_H2_W, fill=FILL_HEADER, align=ALIGN_C)

# 행 30-33: 4공정
procs = ["전극", "조립", "활성화", "출하"]
for i, proc in enumerate(procs):
    rr = 30 + i
    src_row = 25 + i
    # MI Lansing
    set_cell(ws, rr, 1, proc, font=FONT_H2, fill=FILL_LGES, align=ALIGN_C)
    for j, col in enumerate(["C", "D", "E", "F"]):
        set_cell(ws, rr, j + 2, f"='02_MI_Lansing'!{col}{src_row}",
                 font=FONT_CALC, fill=YEAR_FILLS[j], align=ALIGN_C, fmt="0.0%")
    set_cell(ws, rr, 6, f"=E{rr}-B{rr}", font=FONT_DELTA_POS, fill=FILL_SUCCESS, align=ALIGN_C, fmt="+0.0%;-0.0%")
    # GM2
    set_cell(ws, rr, 7, proc, font=FONT_H2, fill=FILL_LGES, align=ALIGN_C)
    for j, col in enumerate(["C", "D", "E", "F"]):
        set_cell(ws, rr, j + 8, f"='04_GM2_각형'!{col}{src_row}",
                 font=FONT_CALC, fill=YEAR_FILLS[j], align=ALIGN_C, fmt="0.0%")
    set_cell(ws, rr, 12, f"=K{rr}-H{rr}", font=FONT_DELTA_POS, fill=FILL_SUCCESS, align=ALIGN_C, fmt="+0.0%;-0.0%")

# 행 34: Total
r = 34
set_cell(ws, r, 1, "Total", font=FONT_TOTAL, fill=FILL_TOTAL, align=ALIGN_C)
for j, col in enumerate(["C", "D", "E", "F"]):
    set_cell(ws, r, j + 2, f"='02_MI_Lansing'!{col}29",
             font=FONT_CALC, fill=YEAR_FILLS[j], align=ALIGN_C, fmt="0.0%")
set_cell(ws, r, 6, "=E34-B34", font=FONT_DELTA_POS, fill=FILL_SUCCESS, align=ALIGN_C, fmt="+0.0%;-0.0%")
set_cell(ws, r, 7, "Total", font=FONT_TOTAL, fill=FILL_TOTAL, align=ALIGN_C)
for j, col in enumerate(["C", "D", "E", "F"]):
    set_cell(ws, r, j + 8, f"='04_GM2_각형'!{col}29",
             font=FONT_CALC, fill=YEAR_FILLS[j], align=ALIGN_C, fmt="0.0%")
set_cell(ws, r, 12, "=K34-H34", font=FONT_DELTA_POS, fill=FILL_SUCCESS, align=ALIGN_C, fmt="+0.0%;-0.0%")

# ───────────────────────────────────────────
# §4. F1 시계열 + 라인 격차 (행 36~44)
# ───────────────────────────────────────────
r = 36
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
set_cell(ws, r, 1, "§4. F1 연도별 사내 사실 + 라인 간 격차", font=FONT_H1, fill=FILL_SUBHEAD, align=ALIGN_L)
ws.row_dimensions[r].height = 22

# 행 37: 헤더
r = 37
set_cell(ws, r, 1, "라인", font=FONT_H2_W, fill=FILL_HEADER, align=ALIGN_C)
set_cell(ws, r, 2, "24년", font=FONT_H2_W, fill=FILL_HEADER, align=ALIGN_C)
for j, y in enumerate(YEAR_LABELS):
    set_cell(ws, r, j + 3, y, font=FONT_H2_W, fill=FILL_HEADER, align=ALIGN_C)
set_cell(ws, r, 7, "추세", font=FONT_H2_W, fill=FILL_HEADER, align=ALIGN_C)
ws.merge_cells(start_row=r, start_column=8, end_row=r, end_column=12)
set_cell(ws, r, 8, "출처 / 비고", font=FONT_H2_W, fill=FILL_HEADER, align=ALIGN_C)

# 행 38: ESMI Lansing
r = 38
set_cell(ws, r, 1, "ESMI Lansing", font=FONT_H2, fill=FILL_F1, align=ALIGN_L)
set_cell(ws, r, 2, "—", font=FONT_NOTE, fill=FILL_F1, align=ALIGN_C)
for j, v in enumerate(F1_TIMESERIES["ESMI Lansing"]):
    set_cell(ws, r, j + 3, v, font=FONT_INPUT, fill=FILL_F1, align=ALIGN_C, fmt="0.0%")
set_cell(ws, r, 7, "=F38-C38", font=FONT_DELTA_POS, fill=FILL_F1, align=ALIGN_C, fmt="+0.0%;-0.0%")
ws.merge_cells(start_row=r, start_column=8, end_row=r, end_column=12)
set_cell(ws, r, 8, "F1 [48~50행] 사내 사실", font=FONT_LINK, fill=FILL_F1, align=ALIGN_L)

# 행 39: ESGM2 각형
r = 39
set_cell(ws, r, 1, "ESGM2 각형", font=FONT_H2, fill=FILL_F1, align=ALIGN_L)
set_cell(ws, r, 2, "—", font=FONT_NOTE, fill=FILL_F1, align=ALIGN_C)
for j, v in enumerate(F1_TIMESERIES["ESGM2 각형"]):
    set_cell(ws, r, j + 3, v, font=FONT_INPUT, fill=FILL_F1, align=ALIGN_C, fmt="0.0%")
set_cell(ws, r, 7, "=F39-C39", font=FONT_DELTA_POS, fill=FILL_F1, align=ALIGN_C, fmt="+0.0%;-0.0%")
ws.merge_cells(start_row=r, start_column=8, end_row=r, end_column=12)
set_cell(ws, r, 8, "F1 [69~73행] 사내 사실", font=FONT_LINK, fill=FILL_F1, align=ALIGN_L)

# 행 40: ESHG (24년 사실 + '25~'28 추정)
r = 40
set_cell(ws, r, 1, "ESHG", font=FONT_H2, fill=FILL_F1, align=ALIGN_L)
set_cell(ws, r, 2, 0.24, font=FONT_INPUT, fill=FILL_F1, align=ALIGN_C, fmt="0.0%")
for j, k in enumerate(["'25", "'26", "'27", "'28"]):
    set_cell(ws, r, j + 3, HG_ESTIMATE[k], font=FONT_INPUT, fill=FILL_F1, align=ALIGN_C, fmt="0.0%")
set_cell(ws, r, 7, "=F40-C40", font=FONT_DELTA_POS, fill=FILL_F1, align=ALIGN_C, fmt="+0.0%;-0.0%")
ws.merge_cells(start_row=r, start_column=8, end_row=r, end_column=12)
set_cell(ws, r, 8, "F1 / S1 [42행] (24년 사실), '25~'28 추정", font=FONT_TAG_WARN, fill=FILL_F1, align=ALIGN_L)

# 행 42: 격차 헤더
r = 42
set_cell(ws, r, 1, "라인 격차", font=FONT_H2_W, fill=FILL_HEADER, align=ALIGN_C)
set_cell(ws, r, 2, "비교", font=FONT_H2_W, fill=FILL_HEADER, align=ALIGN_C)
for j, y in enumerate(YEAR_LABELS):
    set_cell(ws, r, j + 3, y, font=FONT_H2_W, fill=FILL_HEADER, align=ALIGN_C)
ws.merge_cells(start_row=r, start_column=7, end_row=r, end_column=12)
set_cell(ws, r, 7, "해석", font=FONT_H2_W, fill=FILL_HEADER, align=ALIGN_C)

gaps_data = [
    ("vs1", "GM2 - MI Lansing", ["=B15-B14", "=C15-C14", "=D15-D14", "=E15-E14"],
     "GM2 각형이 우위 (전극 모바일셔틀 100% / 조립 SMC 확대)"),
    ("vs2", "MI Lansing - HG", ["=B14-B16", "=C14-C16", "=D14-D16", "=E14-E16"],
     "MI Lansing 우위 (출하·활성화 SMC 효과, HG 활성화 Aging 잔존)"),
    ("vs3", "GM2 - HG", ["=B15-B16", "=C15-C16", "=D15-D16", "=E15-E16"],
     "각형 라인이 HG 대비 최대 격차 (전극 100% 전환 효과)"),
]
for i, (vk, label, fs, note) in enumerate(gaps_data):
    rr = 43 + i
    set_cell(ws, rr, 1, vk, font=FONT_NOTE, fill=FILL_SUBHEAD, align=ALIGN_C)
    set_cell(ws, rr, 2, label, font=FONT_BODY, fill=FILL_S1, align=ALIGN_L)
    for j, f in enumerate(fs):
        set_cell(ws, rr, j + 3, f, font=FONT_DELTA_POS, fill=YEAR_FILLS[j], align=ALIGN_C, fmt="+0.0%;-0.0%")
    ws.merge_cells(start_row=rr, start_column=7, end_row=rr, end_column=12)
    set_cell(ws, rr, 7, note, font=FONT_NOTE, fill=FILL_S1, align=ALIGN_L)

# ───────────────────────────────────────────
# §5. 공정별 적용 vs 잔여 (★ 신규, 행 48~58)
# ───────────────────────────────────────────
r = 48
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
set_cell(ws, r, 1, "§5. 공정별 적용 과제 vs 잔여 (28년 시점)   ★ 어떤 과제가 적용됐고 무엇이 남았는지",
         font=FONT_H1, fill=FILL_SUBHEAD, align=ALIGN_L)
ws.row_dimensions[r].height = 22

# 행 49: 헤더
r = 49
set_cell(ws, r, 1, "공정", font=FONT_H2_W, fill=FILL_HEADER, align=ALIGN_C)
set_cell(ws, r, 2, "적용 연도", font=FONT_H2_W, fill=FILL_HEADER, align=ALIGN_C)
ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
set_cell(ws, r, 3, "✓ 적용된 과제 (MI Lansing + GM2)", font=FONT_H2_W, fill=FILL_HEADER, align=ALIGN_L)
ws.merge_cells(start_row=r, start_column=8, end_row=r, end_column=12)
set_cell(ws, r, 8, "△ 잔여 (28년 시점 미해결)", font=FONT_H2_W, fill=FILL_HEADER, align=ALIGN_L)

applied_vs_remaining = [
    {
        "proc": "전극",
        "year": "'26",
        "applied": "점보롤 스토커 Rack → SMC (MI L) · 모바일셔틀 Rack → SMC (GM2) · 브릿지 물류 → SMC · 밀폐형 AMR",
        "applied_state": "✓ 100% 전환 완료",
        "remaining": "(잔여 없음)",
        "remaining_state": "완료",
        "is_complete": True,
    },
    {
        "proc": "조립",
        "year": "'25~'28 점진",
        "applied": "팬케익 Rack → SMC (MI L) · 자재창고 Rack → SMC (GM2)",
        "applied_state": "✓ 부분 적용",
        "remaining": "PKG 배출 CNV · 조립 원부자재 OHMS (MI L) · OHMS Rail · Lift · 포장해체기 (GM2)",
        "remaining_state": "△ 잔여",
        "is_complete": False,
    },
    {
        "proc": "활성화",
        "year": "'27",
        "applied": "활성화 Rack → SMC · MMF (활성화 SMC) 적용",
        "applied_state": "✓ 부분 적용",
        "remaining": "Aging Lift · 연결 물류 (MI L) · Encloser · HVC/충방전기 Stocker · 연결 물류 (GM2)",
        "remaining_state": "△ 잔여",
        "is_complete": False,
    },
    {
        "proc": "출하",
        "year": "'28",
        "applied": "출하 Rack → SMC · MMF · 밀폐형 AMR",
        "applied_state": "✓ 부분 적용",
        "remaining": "포장기 · 세척기 (MI L+GM2) · 연결물류 (GM2)",
        "remaining_state": "△ 잔여",
        "is_complete": False,
    },
]

for i, row in enumerate(applied_vs_remaining):
    rr = 50 + i
    set_cell(ws, rr, 1, row["proc"], font=FONT_H2, fill=FILL_LGES, align=ALIGN_C)
    set_cell(ws, rr, 2, row["year"], font=FONT_INPUT, fill=YEAR_FILLS[["'25","'26","'27","'28"].index(row["year"]) if row["year"] in ["'25","'26","'27","'28"] else 0], align=ALIGN_C)
    ws.merge_cells(start_row=rr, start_column=3, end_row=rr, end_column=7)
    fill_app = FILL_SUCCESS if row["is_complete"] else FILL_S1
    set_cell(ws, rr, 3, row["applied"], font=FONT_BODY, fill=fill_app, align=ALIGN_L)
    ws.merge_cells(start_row=rr, start_column=8, end_row=rr, end_column=12)
    fill_rem = FILL_SUCCESS if row["is_complete"] else FILL_WARN
    font_rem = FONT_DELTA_POS if row["is_complete"] else FONT_TAG_WARN
    set_cell(ws, rr, 8, row["remaining"], font=font_rem, fill=fill_rem, align=ALIGN_L)
    ws.row_dimensions[rr].height = 36

# 행 55: 요약 footer
r = 55
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
set_cell(ws, r, 1,
         "  ✓ 전극은 '26년 SMC 적용으로 100% 유연 전환 완료.  △ 조립·활성화·출하는 28년 시점에도 잔여 (포장기·세척기·Aging Lift·OHMS 등) 존재 → 후속 과제 발굴 필요",
         font=FONT_BODY, fill=FILL_KEY, align=Alignment(horizontal="left", vertical="center"))
ws.row_dimensions[r].height = 26

# ───────────────────────────────────────────
# §6. 출처 (행 57~)
# ───────────────────────────────────────────
r = 57
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
set_cell(ws, r, 1, "§6. 출처", font=FONT_H1, fill=FILL_SUBHEAD, align=ALIGN_L)
sources_d = [
    ("S1", "references/FA설비유연율 Data.md (393줄, 4 시트 raw)"),
    ("F1", "references/FA 유연성 지표 정의 (사내 사실, 2026-05-21)"),
    ("F3", "references/26FA KPI.md"),
    ("F4", "references/roadmap/2026_FA기술담당_중장기로드맵_v2.md"),
    ("내부", "본 대시보드는 시트 02·04·06 의 셀을 수식으로 참조. 입력값 변경 시 즉시 갱신."),
    ("추정", "HG '25~'28 (행 16·23·40) 은 24년 24% 실측 기반 추정 — 활성화 Aging 잔존 가정"),
]
for i, (k, v) in enumerate(sources_d):
    set_cell(ws, 58 + i, 1, k, font=FONT_H2_W, fill=FILL_HEADER, align=ALIGN_C)
    ws.merge_cells(start_row=58+i, start_column=2, end_row=58+i, end_column=12)
    set_cell(ws, 58 + i, 2, v, font=FONT_LINK, fill=FILL_S1, align=ALIGN_L)



def build_line_sheet(ws, line_name, line_short, data, yearly, source_rows, line_meta, ts_key):
    widths = [12, 12, 14, 14, 14, 14, 30, 30]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    title_band(ws, 1, f"{line_name} — 유연율 표준화 + 연도별 매트릭스", 8)

    # 라인 메타
    set_cell(ws, 3, 1, "라인", font=FONT_H2, fill=FILL_HEADER, align=ALIGN_L)
    ws.merge_cells("B3:H3")
    set_cell(ws, 3, 2, line_meta, font=FONT_BODY, align=ALIGN_L)
    set_cell(ws, 4, 1, "출처", font=FONT_H2, fill=FILL_HEADER, align=ALIGN_L)
    ws.merge_cells("B4:H4")
    set_cell(ws, 4, 2, f"S1 [{source_rows}]", font=FONT_LINK, align=ALIGN_L)

    # §1 시나리오 1 ('25 현재)
    r = 6
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    set_cell(ws, r, 1, "§1. 시나리오 1 — '25 현재 (S1 좌측, 4공정 라인 합계)", font=FONT_H1, fill=FILL_S1, align=ALIGN_L)
    r = 7
    headers = ["공정", "구분", "Total (억원)", "유연설비", "고정설비", "유연율", "비고", "수식 (참고)"]
    for j, h in enumerate(headers):
        set_cell(ws, r, j + 1, h, font=FONT_H2, fill=FILL_HEADER, align=ALIGN_C)

    for i, row in enumerate(data):
        name, t1, f1, x1, n1, t2, f2, x2, n2 = row
        rr = 8 + i
        set_cell(ws, rr, 1, name, font=FONT_H2, fill=FILL_LGES, align=ALIGN_C)
        set_cell(ws, rr, 2, line_short, font=FONT_NOTE, fill=FILL_LGES, align=ALIGN_C)
        set_cell(ws, rr, 3, t1, font=FONT_INPUT, fill=FILL_S1, align=ALIGN_R, fmt="#,##0.0")
        set_cell(ws, rr, 4, f1, font=FONT_INPUT, fill=FILL_S1, align=ALIGN_R, fmt="#,##0.0")
        set_cell(ws, rr, 5, x1, font=FONT_INPUT, fill=FILL_S1, align=ALIGN_R, fmt="#,##0.0")
        set_cell(ws, rr, 6, f"=D{rr}/C{rr}", font=FONT_CALC, fill=FILL_S1, align=ALIGN_C, fmt="0.0%")
        set_cell(ws, rr, 7, n1, font=FONT_NOTE, fill=FILL_S1, align=ALIGN_L)
        set_cell(ws, rr, 8, f"=D{rr}/C{rr}", font=FONT_FORMULA, fill=FILL_S1, align=ALIGN_L)

    r = 12
    set_cell(ws, r, 1, "Total", font=FONT_TOTAL, fill=FILL_TOTAL, align=ALIGN_C)
    set_cell(ws, r, 2, line_short, font=FONT_NOTE, fill=FILL_TOTAL, align=ALIGN_C)
    set_cell(ws, r, 3, "=SUM(C8:C11)", font=FONT_CALC, fill=FILL_S1, align=ALIGN_R, fmt="#,##0.0")
    set_cell(ws, r, 4, "=SUM(D8:D11)", font=FONT_CALC, fill=FILL_S1, align=ALIGN_R, fmt="#,##0.0")
    set_cell(ws, r, 5, "=SUM(E8:E11)", font=FONT_CALC, fill=FILL_S1, align=ALIGN_R, fmt="#,##0.0")
    set_cell(ws, r, 6, "=D12/C12", font=FONT_CALC, fill=FILL_KEY, align=ALIGN_C, fmt="0.0%")
    set_cell(ws, r, 7, "라인 합계", font=FONT_NOTE, fill=FILL_S1, align=ALIGN_L)
    set_cell(ws, r, 8, "=SUM(D8:D11)/SUM(C8:C11)", font=FONT_FORMULA, fill=FILL_S1, align=ALIGN_L)

    # §2 시나리오 2 ('28)
    r = 14
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    set_cell(ws, r, 1, "§2. 시나리오 2 — '28 신규 과제 적용 후 (S1 우측, 28년 컬럼)", font=FONT_H1, fill=FILL_S2, align=ALIGN_L)
    r = 15
    for j, h in enumerate(headers):
        set_cell(ws, r, j + 1, h, font=FONT_H2, fill=FILL_HEADER, align=ALIGN_C)
    for i, row in enumerate(data):
        name, t1, f1, x1, n1, t2, f2, x2, n2 = row
        rr = 16 + i
        set_cell(ws, rr, 1, name, font=FONT_H2, fill=FILL_LGES, align=ALIGN_C)
        set_cell(ws, rr, 2, line_short, font=FONT_NOTE, fill=FILL_LGES, align=ALIGN_C)
        set_cell(ws, rr, 3, t2, font=FONT_INPUT, fill=FILL_S2, align=ALIGN_R, fmt="#,##0.0")
        set_cell(ws, rr, 4, f2, font=FONT_INPUT, fill=FILL_S2, align=ALIGN_R, fmt="#,##0.0")
        set_cell(ws, rr, 5, x2, font=FONT_INPUT, fill=FILL_S2, align=ALIGN_R, fmt="#,##0.0")
        set_cell(ws, rr, 6, f"=D{rr}/C{rr}", font=FONT_CALC, fill=FILL_S2, align=ALIGN_C, fmt="0.0%")
        set_cell(ws, rr, 7, n2, font=FONT_NOTE, fill=FILL_S2, align=ALIGN_L)
        set_cell(ws, rr, 8, f"=D{rr}/C{rr}", font=FONT_FORMULA, fill=FILL_S2, align=ALIGN_L)
    r = 20
    set_cell(ws, r, 1, "Total", font=FONT_TOTAL, fill=FILL_TOTAL, align=ALIGN_C)
    set_cell(ws, r, 2, line_short, font=FONT_NOTE, fill=FILL_TOTAL, align=ALIGN_C)
    set_cell(ws, r, 3, "=SUM(C16:C19)", font=FONT_CALC, fill=FILL_S2, align=ALIGN_R, fmt="#,##0.0")
    set_cell(ws, r, 4, "=SUM(D16:D19)", font=FONT_CALC, fill=FILL_S2, align=ALIGN_R, fmt="#,##0.0")
    set_cell(ws, r, 5, "=SUM(E16:E19)", font=FONT_CALC, fill=FILL_S2, align=ALIGN_R, fmt="#,##0.0")
    set_cell(ws, r, 6, "=D20/C20", font=FONT_CALC, fill=FILL_KEY, align=ALIGN_C, fmt="0.0%")
    set_cell(ws, r, 7, "라인 합계", font=FONT_NOTE, fill=FILL_S2, align=ALIGN_L)
    set_cell(ws, r, 8, "=SUM(D16:D19)/SUM(C16:C19)", font=FONT_FORMULA, fill=FILL_S2, align=ALIGN_L)

    # §3 연도별 매트릭스 (★ 신규)
    r = 22
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    set_cell(ws, r, 1, "§3. 연도별 유연율 매트릭스 (★ 세로: 공정 / 가로: 연도 / 각 셀: 누적 과제 적용 후 유연율)",
             font=FONT_H1, fill=FILL_KEY, align=ALIGN_L)

    # 헤더 — 연도별 과제 표시
    r = 23
    set_cell(ws, r, 1, "공정", font=FONT_H2, fill=FILL_HEADER, align=ALIGN_C)
    set_cell(ws, r, 2, "Total ('25)", font=FONT_H2, fill=FILL_HEADER, align=ALIGN_C)
    year_fills = [FILL_Y25, FILL_Y26, FILL_Y27, FILL_Y28]
    for j, y in enumerate(["'25년", "'26년", "'27년", "'28년"]):
        set_cell(ws, r, j + 3, y, font=FONT_H2, fill=year_fills[j], align=ALIGN_C)
    set_cell(ws, r, 7, "공정별 적용 과제", font=FONT_H2, fill=FILL_HEADER, align=ALIGN_C)
    set_cell(ws, r, 8, "개선폭 ('25→'28)", font=FONT_H2, fill=FILL_HEADER, align=ALIGN_C)

    # 과제 행 — 연도별
    r = 24
    set_cell(ws, r, 1, "주요 과제", font=FONT_H2, fill=FILL_HEADER, align=ALIGN_L)
    set_cell(ws, r, 2, "—", font=FONT_NOTE, fill=FILL_HEADER, align=ALIGN_C)
    for j, k in enumerate(["'25", "'26", "'27", "'28"]):
        set_cell(ws, r, j + 3, F1_TASKS[k], font=FONT_BODY, fill=year_fills[j], align=ALIGN_C)
    ws.merge_cells("G24:H24")
    set_cell(ws, r, 7, "출처: S1 [44행]", font=FONT_LINK, align=ALIGN_L)

    # 공정별 행 (4공정)
    for i, (proc, vals) in enumerate(yearly.items()):
        rr = 25 + i
        set_cell(ws, rr, 1, proc, font=FONT_H2, fill=FILL_LGES, align=ALIGN_C)
        # Total ('25) 셀 참조 (시나리오 1 합계)
        set_cell(ws, rr, 2, f"=C{8+i}", font=FONT_CALC, fill=FILL_LGES, align=ALIGN_R, fmt="#,##0.0")
        # 연도별 유연율 (입력값)
        for j, v in enumerate(vals):
            set_cell(ws, rr, j + 3, v, font=FONT_INPUT, fill=year_fills[j],
                     align=ALIGN_C, fmt="0.0%")
        # 비고
        set_cell(ws, rr, 7, TASK_BY_PROCESS[proc], font=FONT_NOTE, align=ALIGN_L)
        # 개선폭
        set_cell(ws, rr, 8, f"=F{rr}-C{rr}", font=FONT_CALC, fill=FILL_KEY,
                 align=ALIGN_C, fmt="+0.0%;-0.0%")

    # Total 행 (SUMPRODUCT 가중평균)
    r = 29
    set_cell(ws, r, 1, "라인 Total", font=FONT_TOTAL, fill=FILL_TOTAL, align=ALIGN_C)
    set_cell(ws, r, 2, "=SUM(B25:B28)", font=FONT_CALC, fill=FILL_TOTAL, align=ALIGN_R, fmt="#,##0.0")
    # 연도별 가중평균 = SUMPRODUCT(공정별 Total, 공정별 유연율) / SUM(공정별 Total)
    for j in range(4):
        col_y = get_column_letter(j + 3)
        formula = f"=SUMPRODUCT(B25:B28,{col_y}25:{col_y}28)/SUM(B25:B28)"
        set_cell(ws, r, j + 3, formula, font=FONT_CALC, fill=FILL_KEY,
                 align=ALIGN_C, fmt="0.0%")
    set_cell(ws, r, 7, "= SUMPRODUCT(공정 Total, 공정 유연율) / SUM(공정 Total)",
             font=FONT_FORMULA, align=ALIGN_L)
    set_cell(ws, r, 8, "=F29-C29", font=FONT_CALC, fill=FILL_KEY,
             align=ALIGN_C, fmt="+0.0%;-0.0%")

    # F1 시계열 비교 (검증)
    r = 31
    set_cell(ws, r, 1, "F1 사실", font=FONT_H2, fill=FILL_F1, align=ALIGN_C)
    set_cell(ws, r, 2, "—", font=FONT_NOTE, fill=FILL_F1, align=ALIGN_C)
    for j, v in enumerate(F1_TIMESERIES[ts_key]):
        set_cell(ws, r, j + 3, v, font=FONT_INPUT, fill=FILL_F1, align=ALIGN_C, fmt="0.0%")
    ws.merge_cells("G31:H31")
    set_cell(ws, r, 7, "F1 [48~50, 69~73] / 본 시트 매트릭스 결과와 ±2%p 이내 정합 검증",
             font=FONT_LINK, align=ALIGN_L)

    # 검증 박스
    r = 33
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    set_cell(ws, r, 1,
             f"★ 본 시트 매트릭스 = '25 시나리오 1 + 연도별 과제 누적 적용 모델. 라인 Total 의 연도별 결과가 F1 시계열과 ±2%p 이내 정합 시 모델 OK.",
             font=FONT_NOTE, fill=FILL_KEY, align=ALIGN_L)

    # §4 공정별 적용 과제 상세
    r = 35
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    set_cell(ws, r, 1, "§4. 공정별 적용 과제 상세 (라인별 동일)", font=FONT_H1, fill=FILL_HEADER, align=ALIGN_L)
    r = 36
    set_cell(ws, r, 1, "공정", font=FONT_H2, fill=FILL_HEADER, align=ALIGN_C)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    set_cell(ws, r, 2, "적용 과제", font=FONT_H2, fill=FILL_HEADER, align=ALIGN_C)
    set_cell(ws, r, 5, "적용 연도", font=FONT_H2, fill=FILL_HEADER, align=ALIGN_C)
    ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=8)
    set_cell(ws, r, 6, "주요 변화 설명", font=FONT_H2, fill=FILL_HEADER, align=ALIGN_C)

    proc_tasks = [
        ("전극", "전극 SMC (모바일셔틀/점보롤 Rack → SMC 전환)", "'26", "고정형 → 유연형 100% 전환"),
        ("조립", "팬케익 Rack → SMC, PKG 배출 CNV/OHMS 점진 개선", "'25~'28", "단계별 SMC 적용 + 잔여 고정형 정리"),
        ("활성화", "MMF (활성화 Rack → SMC, 활성화 SMC 적용)", "'27", "Aging Lift 외 활성화 영역 SMC 전환"),
        ("출하", "밀폐형 AMR (출하 Rack → SMC, 밀폐형 AMR)", "'28", "포장기/세척기 외 출하 영역 AMR 전환"),
    ]
    for i, (p, t, y, n) in enumerate(proc_tasks):
        rr = 37 + i
        set_cell(ws, rr, 1, p, font=FONT_H2, fill=FILL_LGES, align=ALIGN_C)
        ws.merge_cells(start_row=rr, start_column=2, end_row=rr, end_column=4)
        set_cell(ws, rr, 2, t, font=FONT_BODY, align=ALIGN_L)
        set_cell(ws, rr, 5, y, font=FONT_H2, fill=FILL_KEY, align=ALIGN_C)
        ws.merge_cells(start_row=rr, start_column=6, end_row=rr, end_column=8)
        set_cell(ws, rr, 6, n, font=FONT_NOTE, align=ALIGN_L)




# ============================================================
# Sheet 2 — MI Lansing
# ============================================================
ws = wb.create_sheet("02_MI_Lansing")
build_line_sheet(ws, "MI Lansing (ESMI L)", "MI L", MI_LANSING, MI_LANSING_YEARLY,
                 "46~107행 (raw) + 24~28행 (서머리)",
                 "LGES 미국 미시간 Lansing 각형 라인 / 1차 양산 가동 / 4공정: 전극·조립·활성화·출하",
                 "ESMI Lansing")

# Sheet 3 — MI Lansing raw
ws = wb.create_sheet("03_MI_Lansing_raw")
rows = parse_md_section(md_text, "유연율 정의_ESMI L_세분화_10%절감")
write_raw_sheet(ws, "MI Lansing — S1 raw 표 (유연율 정의_ESMI L_세분화_10%절감)", rows, "46~107행")

# Sheet 4 — GM2
ws = wb.create_sheet("04_GM2_각형")
build_line_sheet(ws, "GM2 각형 (ESGM2)", "GM2", GM2, GM2_YEARLY,
                 "108~321행 (raw) + 30~34행 (서머리)",
                 "LGES GM2 각형 라인 / 신규 / 4공정: 전극·조립·활성화·출하",
                 "ESGM2 각형")

# Sheet 5 — GM2 raw
ws = wb.create_sheet("05_GM2_raw")
rows = parse_md_section(md_text, "유연율 검토_ESGM2_10%절감")
write_raw_sheet(ws, "GM2 각형 — S1 raw 표 (유연율 검토_ESGM2_10%절감)", rows, "108~321행")

# Sheet 6 — HG
ws = wb.create_sheet("06_HG")
widths = [16, 16, 16, 16, 16, 16, 30]
for i, w in enumerate(widths):
    ws.column_dimensions[get_column_letter(i + 1)].width = w
title_band(ws, 1, "HG (ESHG) — 유연율 표준화 (합계만, 4공정 분해 미공시)", 7)
set_cell(ws, 3, 1, "라인", font=FONT_H2, fill=FILL_HEADER, align=ALIGN_L)
ws.merge_cells("B3:G3")
set_cell(ws, 3, 2, "LGES HG 라인 / S1 raw 시트는 설비 단위(AMR·모바일셔틀·OHT·Aging·포장기 등) 로 분해. 본 시트는 합계만 박제. 디테일은 07_HG_raw 참조.", font=FONT_BODY, align=ALIGN_L)
set_cell(ws, 4, 1, "출처", font=FONT_H2, fill=FILL_HEADER, align=ALIGN_L)
ws.merge_cells("B4:G4")
set_cell(ws, 4, 2, f"S1 [{HG_TOTAL['출처']}] + 서머리 ESHG 24년 = 24%", font=FONT_LINK, align=ALIGN_L)
r = 6
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
set_cell(ws, r, 1, "§1. HG 합계 (24년 기준, 집행품의 기준)", font=FONT_H1, fill=FILL_S1, align=ALIGN_L)
r = 7
hg_headers = ["라인", "예산 총액 (억원)", "Total (= 설비비+설치비)", "유연설비", "고정설비", "유연율", "비고"]
for j, h in enumerate(hg_headers):
    set_cell(ws, r, j + 1, h, font=FONT_H2, fill=FILL_HEADER, align=ALIGN_C)
r = 8
set_cell(ws, r, 1, "HG", font=FONT_H2, fill=FILL_LGES, align=ALIGN_C)
set_cell(ws, r, 2, HG_TOTAL["예산 총액"], font=FONT_INPUT, fill=FILL_S1, align=ALIGN_R, fmt="#,##0.0")
set_cell(ws, r, 3, "=D8+E8", font=FONT_CALC, fill=FILL_S1, align=ALIGN_R, fmt="#,##0.0")
set_cell(ws, r, 4, HG_TOTAL["유연설비"], font=FONT_INPUT, fill=FILL_S1, align=ALIGN_R, fmt="#,##0.0")
set_cell(ws, r, 5, HG_TOTAL["고정설비"], font=FONT_INPUT, fill=FILL_S1, align=ALIGN_R, fmt="#,##0.0")
set_cell(ws, r, 6, "=D8/C8", font=FONT_CALC, fill=FILL_KEY, align=ALIGN_C, fmt="0.0%")
set_cell(ws, r, 7, HG_TOTAL["비고"], font=FONT_NOTE, align=ALIGN_L)

r = 10
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
set_cell(ws, r, 1, "§2. HG 투자비 Breakdown", font=FONT_H1, fill=FILL_HEADER, align=ALIGN_L)
r = 11
for j, h in enumerate(["구분", "금액 (억원)", "비중", "비고", "", "", ""]):
    set_cell(ws, r, j + 1, h, font=FONT_H2, fill=FILL_HEADER, align=ALIGN_C)
r = 12
set_cell(ws, r, 1, "설비비", font=FONT_BODY, fill=FILL_LGES, align=ALIGN_L)
set_cell(ws, r, 2, HG_TOTAL["설비비"], font=FONT_INPUT, fill=FILL_S1, align=ALIGN_R, fmt="#,##0.0")
set_cell(ws, r, 3, "=B12/(B12+B13)", font=FONT_CALC, fill=FILL_S1, align=ALIGN_C, fmt="0.0%")
ws.merge_cells("D12:G12")
set_cell(ws, r, 4, "S1 [351행]", font=FONT_LINK, align=ALIGN_L)
r = 13
set_cell(ws, r, 1, "설치비", font=FONT_BODY, fill=FILL_LGES, align=ALIGN_L)
set_cell(ws, r, 2, HG_TOTAL["설치비"], font=FONT_INPUT, fill=FILL_S1, align=ALIGN_R, fmt="#,##0.0")
set_cell(ws, r, 3, "=B13/(B12+B13)", font=FONT_CALC, fill=FILL_S1, align=ALIGN_C, fmt="0.0%")
ws.merge_cells("D13:G13")
set_cell(ws, r, 4, "S1 [351행]", font=FONT_LINK, align=ALIGN_L)
r = 14
set_cell(ws, r, 1, "Total", font=FONT_TOTAL, fill=FILL_TOTAL, align=ALIGN_C)
set_cell(ws, r, 2, "=B12+B13", font=FONT_CALC, fill=FILL_KEY, align=ALIGN_R, fmt="#,##0.0")
set_cell(ws, r, 3, "100%", font=FONT_NOTE, fill=FILL_KEY, align=ALIGN_C)
ws.merge_cells("D14:G14")
set_cell(ws, r, 4, "= 설비비 + 설치비", font=FONT_NOTE, align=ALIGN_L)

# 후속 데이터 보강 권고
r = 16
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
set_cell(ws, r, 1, "§3. 후속 데이터 보강 권고", font=FONT_H1, fill=FILL_HEADER, align=ALIGN_L)
recs = [
    "07_HG_raw 시트의 설비별 데이터를 4공정(전극·조립·활성화·출하·기타)으로 재집계하여 본 시트 §1 확장 → MI Lansing/GM2 와 동일 표준 구조 달성",
    "'25~'28 ESHG 시계열 ('26·'27·'28 미공시) 보강 — F1 갱신 필요",
    "HG '28 시나리오(신규 과제 적용 후) 추정치 추가 — S1 우측 컬럼 부재로 본 시트 미작성",
]
for i, rec in enumerate(recs):
    rr = 17 + i
    set_cell(ws, rr, 1, f"권고 {i+1}", font=FONT_H2, fill=FILL_HEADER, align=ALIGN_C)
    ws.merge_cells(start_row=rr, start_column=2, end_row=rr, end_column=7)
    set_cell(ws, rr, 2, rec, font=FONT_BODY, align=ALIGN_L)

# Sheet 7 — HG raw
ws = wb.create_sheet("07_HG_raw")
rows = parse_md_section(md_text, "유연율 검토_ESHG_세분화")
write_raw_sheet(ws, "HG — S1 raw 표 (유연율 검토_ESHG_세분화)", rows, "322~393행")


# ============================================================
# 저장
# ============================================================
OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)

size_kb = OUT.stat().st_size / 1024
print(f"saved -> {OUT}")
print(f"size  -> {size_kb:.1f} KB")
print(f"sheets -> {len(wb.sheetnames)} ({', '.join(wb.sheetnames)})")

from openpyxl import load_workbook
wb2 = load_workbook(OUT)
total_f = 0
for sn in wb2.sheetnames:
    for row in wb2[sn].iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value.startswith("="):
                total_f += 1
print(f"formulas -> {total_f} 개")

# 정합성 검증
def verify_matrix(yearly, line_data, label, f1):
    totals = [d[1] for d in line_data]  # '25 Total
    yearly_results = []
    for y in range(4):
        flex_by_proc = [yearly[d[0]][y] for d in line_data]
        result = sum(t * f for t, f in zip(totals, flex_by_proc)) / sum(totals)
        yearly_results.append(result)
    print(f"\n=== {label} 연도별 매트릭스 검증 ===")
    for y, label_y in enumerate(["'25", "'26", "'27", "'28"]):
        diff = (yearly_results[y] - f1[y]) * 100 if f1[y] else 0
        print(f"  {label_y}: 매트릭스 {yearly_results[y]*100:.1f}%  vs F1 {f1[y]*100:.0f}%  ({diff:+.1f}%p)")

verify_matrix(MI_LANSING_YEARLY, MI_LANSING, "MI Lansing", F1_TIMESERIES["ESMI Lansing"])
verify_matrix(GM2_YEARLY, GM2, "GM2 각형", F1_TIMESERIES["ESGM2 각형"])
