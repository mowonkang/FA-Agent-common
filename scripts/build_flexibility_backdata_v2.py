"""
유연성 지표 정량 백데이터 (v2 기준) — xlsx 빌더 v2 (계산 함수 노출)

★ v1 (이전) 대비 변경
- 시트 03·04 (SDI/CATL 계산) 에 SUMPRODUCT 함수 도입
- 공장별 자체 유연성 × 공장별 GWh 가중치 = Base 시나리오 자동 계산
- 시트 05 의 SDI/CATL Base 참조 행 번호 갱신

- 출력: outputs/유연성_경쟁사_정량백데이터_v2_2026-05-21.xlsx
- 구조: 10 sheets
- 핵심 원칙
  · LGES 셀: F1 (사내 사실) — 하드코딩 (변경 불가)
  · SDI / CATL Base 시나리오 = SUMPRODUCT(공장별 유연성, 공장별 가중치)
  · 보수/낙관 = Base × 조정 계수 (입력값)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

OUT = Path("/home/user/claude_FA-Agent_1/outputs/유연성_경쟁사_정량백데이터_v2_2026-05-21.xlsx")

# ============================================================
# 스타일 정의
# ============================================================
FONT_TITLE = Font(name="맑은 고딕", size=16, bold=True, color="000000")
FONT_H1 = Font(name="맑은 고딕", size=12, bold=True, color="000000")
FONT_H2 = Font(name="맑은 고딕", size=10, bold=True, color="000000")
FONT_BODY = Font(name="맑은 고딕", size=10, color="000000")
FONT_INPUT = Font(name="맑은 고딕", size=10, color="0000FF", bold=True)
FONT_CALC = Font(name="맑은 고딕", size=10, color="000000", bold=True)
FONT_LINK = Font(name="맑은 고딕", size=10, color="006600")
FONT_NOTE = Font(name="맑은 고딕", size=9, italic=True, color="595959")
FONT_FOOT = Font(name="맑은 고딕", size=8, bold=True, color="006600")
FONT_FORMULA = Font(name="Consolas", size=9, italic=True, color="C00000")

FILL_LGES = PatternFill("solid", fgColor="E2EFD9")
FILL_SDI = PatternFill("solid", fgColor="FFF2CC")
FILL_CATL = PatternFill("solid", fgColor="FCE4D6")
FILL_KEY = PatternFill("solid", fgColor="FFFF00")
FILL_HEADER = PatternFill("solid", fgColor="BFBFBF")
FILL_TITLE = PatternFill("solid", fgColor="404040")
FILL_BASE = PatternFill("solid", fgColor="DEEBF7")
FONT_TITLE_W = Font(name="맑은 고딕", size=14, bold=True, color="FFFFFF")

THIN = Side(style="thin", color="808080")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
ALIGN_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_L = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_R = Alignment(horizontal="right", vertical="center", wrap_text=True)


def set_cell(ws, row, col, value, font=None, fill=None, align=None, border=BOX, fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    if font: c.font = font
    if fill: c.fill = fill
    if align: c.alignment = align
    if border: c.border = border
    if fmt: c.number_format = fmt
    return c


def add_title_band(ws, row, text, span):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = FONT_TITLE_W
    c.fill = FILL_TITLE
    c.alignment = ALIGN_C
    ws.row_dimensions[row].height = 24


# ============================================================
wb = Workbook()
wb.remove(wb.active)


# ============================================================
# Sheet 00 — Cover
# ============================================================
ws = wb.create_sheet("00_Cover")
ws.column_dimensions["A"].width = 22
ws.column_dimensions["B"].width = 70

add_title_band(ws, 1, "LGES 유연성 지표 — 경쟁사 정량 백데이터 (v2, 계산 함수 노출)", 4)
ws.row_dimensions[1].height = 28

cover_meta = [
    ("보고 일자", "2026-05-21"),
    ("작성", "FA기술담당 (Claude AI 분석 에이전트 생성)"),
    ("버전", "v2 (최종 결정판) — SUMPRODUCT 함수로 계산 로직 노출"),
    ("정의 기준", "F1 (references/FA 유연성 지표 정의, 2026-05-21, 기계장치 70:30)"),
    ("산식", "Flexibility = Σ(유연 설비 투자비) / Σ(전체 설비 투자비), 분자·분모 모두 (설비비 + 셋업비) 합계"),
    ("유연 설비", "셋업비 비중 30% 이하 (AMR, AGV, OHT, Shuttle, SMC, Skid/대차)"),
    ("고정 설비", "셋업비 30% 초과 (Stacker Crane, Conveyor, Lift, 세척기, 포장기, MSS Rack, 구조물)"),
    ("기간", "'25년 ~ '28년 (4개년)"),
    ("용도", "CEO 보고용 정량 백데이터 (시나리오 변경 가능)"),
]
for i, (k, v) in enumerate(cover_meta):
    set_cell(ws, 3 + i, 1, k, font=FONT_H2, fill=FILL_HEADER)
    set_cell(ws, 3 + i, 2, v, font=FONT_BODY, align=ALIGN_L)

set_cell(ws, 13, 1, "시트 구조", font=FONT_H1, fill=FILL_HEADER)
sheets_desc = [
    ("00_Cover", "본 시트 — 보고서 개요 + 색상 범례 + 계산 로직 안내"),
    ("01_정의_분류표", "LGES F1 정의 산식 + 4-row 분류표 + 예외 조항"),
    ("02_LGES_사실", "★ LGES F1 공식 수치 (사내 사실, 변경 불가) + 투자비 Breakdown"),
    ("03_SDI_계산", "★ 7공장 자체 유연성 × GWh 가중치 = SUMPRODUCT Base 시나리오"),
    ("04_CATL_계산", "★ 7공장 자체 유연성 × GWh 가중치 = SUMPRODUCT Base 시나리오"),
    ("05_12셀_매트릭스", "3사 × 4개년 통합표 (시트 02·03·04 자동 참조)"),
    ("06_가정카드", "A1~A10 가정 + 영향 셀 + 근거 출처 매핑"),
    ("07_공정별_매트릭스", "전극·조립·화성·팩·물류 × 3사 자동화 성숙도"),
    ("08_라인재구성_비용", "화학 변경 / 폼팩터 부분 / 신규 폼팩터 × 3사"),
    ("09_출처", "F1~F5 + D1~D10 + E1~E15 = 총 30개 출처"),
]
for i, (sn, desc) in enumerate(sheets_desc):
    set_cell(ws, 14 + i, 1, sn, font=FONT_LINK, align=ALIGN_L)
    set_cell(ws, 14 + i, 2, desc, font=FONT_BODY, align=ALIGN_L)

set_cell(ws, 26, 1, "★ 계산 로직 요약", font=FONT_H1, fill=FILL_KEY)
ws.merge_cells("A26:B26")
calc_logic = [
    ("LGES 사실", "MI Lansing 47/61/75/76% · ESGM2 50/62/74/78% — F1 사내 공식 (시트 02), 변경 불가"),
    ("SDI Base 계산", "= SUMPRODUCT(7공장 '25년 유연성, 7공장 '25년 GWh 가중치) — 시트 03 (연도별 4개 수식)"),
    ("CATL Base 계산", "= SUMPRODUCT(7공장 '25년 유연성, 7공장 '25년 GWh 가중치) — 시트 04 (연도별 4개 수식)"),
    ("보수 시나리오", "= Base × 보수 계수 (입력값, 기본 0.88~0.91)"),
    ("낙관 시나리오", "= Base × 낙관 계수 (입력값, 기본 1.09~1.14)"),
    ("12셀 매트릭스 LGES 행", "= '02_LGES_사실'!B6:E7 (참조 수식)"),
    ("12셀 매트릭스 SDI 행", "= '03_SDI_계산'!B36:E36 (Base 시나리오 참조)"),
    ("12셀 매트릭스 CATL 행", "= '04_CATL_계산'!B36:E36 (Base 시나리오 참조)"),
    ("LGES vs SDI/CATL 격차", "= LGES MI Lansing − SDI/CATL Base (자동 계산)"),
]
for i, (k, v) in enumerate(calc_logic):
    set_cell(ws, 27 + i, 1, k, font=FONT_H2, fill=FILL_HEADER)
    set_cell(ws, 27 + i, 2, v, font=FONT_BODY, align=ALIGN_L)

set_cell(ws, 38, 1, "색상 범례", font=FONT_H1, fill=FILL_HEADER)
legends = [
    ("파란색 텍스트", "하드코딩 입력값 — 시나리오 변경 가능", FONT_INPUT, None),
    ("검은 굵은 글씨", "수식·계산값 (SUMPRODUCT 등) — 자동 갱신", FONT_CALC, None),
    ("빨간 이탤릭", "수식 본문 표시 (참고)", FONT_FORMULA, None),
    ("초록색 텍스트", "시트 링크·출처 코드", FONT_LINK, None),
    ("노란색 배경", "핵심 가정·주의 필요", FONT_BODY, FILL_KEY),
    ("연두색 배경", "LGES 데이터 (사내 사실 F1)", FONT_BODY, FILL_LGES),
    ("연황색 배경", "Samsung SDI 데이터 (AI 추정)", FONT_BODY, FILL_SDI),
    ("연주색 배경", "CATL 데이터 (AI 추정)", FONT_BODY, FILL_CATL),
    ("연하늘 배경", "Base 시나리오 (SUMPRODUCT 결과 행)", FONT_BODY, FILL_BASE),
]
for i, (label, desc, f, fl) in enumerate(legends):
    set_cell(ws, 39 + i, 1, label, font=f, fill=fl, align=ALIGN_L)
    set_cell(ws, 39 + i, 2, desc, font=FONT_BODY, align=ALIGN_L)

set_cell(ws, 49, 1, "출처: outputs/유연성_경쟁사_분석_v2_2026-05-21.md + references/FA 유연성 지표 정의 + references/경쟁사/2026-05-21_*.md (D1~D5)", font=FONT_FOOT, align=ALIGN_L)


# ============================================================
# Sheet 01 — 정의·분류표
# ============================================================
ws = wb.create_sheet("01_정의_분류표")
ws.column_dimensions["A"].width = 14
ws.column_dimensions["B"].width = 22
ws.column_dimensions["C"].width = 14
ws.column_dimensions["D"].width = 42

add_title_band(ws, 1, "F1 유연성 지표 정의 + 분류표 (LGES 사내 사실)", 4)

set_cell(ws, 3, 1, "산식", font=FONT_H2, fill=FILL_HEADER)
ws.merge_cells("B3:D3")
set_cell(ws, 3, 2, "Flexibility = Σ(유연 설비 투자비) / Σ(전체 설비 투자비), 분자·분모 모두 (설비비 + 셋업비) 합계", font=FONT_BODY, align=ALIGN_L)

set_cell(ws, 4, 1, "유연 기준", font=FONT_H2, fill=FILL_HEADER)
ws.merge_cells("B4:D4")
set_cell(ws, 4, 2, "셋업비 비중 30% 이하 → 유연 설비 / 30% 초과 → 고정 설비", font=FONT_BODY, align=ALIGN_L)

set_cell(ws, 5, 1, "출처", font=FONT_H2, fill=FILL_HEADER)
ws.merge_cells("B5:D5")
set_cell(ws, 5, 2, "F1: references/FA 유연성 지표 정의 [14행 산식, 18~21행 기준]", font=FONT_LINK, align=ALIGN_L)

set_cell(ws, 7, 1, "분류표 (4-row)", font=FONT_H1, fill=FILL_HEADER)
ws.merge_cells("A7:D7")
headers = ["구분", "Sub-Group", "제작:설치", "주요 설비"]
for i, h in enumerate(headers):
    set_cell(ws, 8, i + 1, h, font=FONT_H2, fill=FILL_HEADER, align=ALIGN_C)

rows = [
    ("고정 설비", "구조물", "50 : 50", "Rack, Stage 등"),
    ("고정 설비", "기계장치", "70 : 30", "Crane, Conveyor, Lift, 세척기 등"),
    ("유연 설비", "Vehicle (AMR·Shuttle)", "90 : 10", "AGV, AMR, OHT, Shuttle"),
    ("유연 설비", "제작품", "100 : 0", "Skid, 대차"),
]
for i, r in enumerate(rows):
    for j, v in enumerate(r):
        fill = FILL_LGES if "유연" in r[0] else None
        set_cell(ws, 9 + i, j + 1, v, font=FONT_BODY,
                 fill=fill, align=ALIGN_C if j != 3 else ALIGN_L)

set_cell(ws, 13, 1, "출처", font=FONT_H2, fill=FILL_HEADER)
ws.merge_cells("B13:D13")
set_cell(ws, 13, 2, "F1: references/FA 유연성 지표 정의 [32~37행]", font=FONT_LINK, align=ALIGN_L)

set_cell(ws, 15, 1, "예외 조항", font=FONT_H1, fill=FILL_HEADER)
ws.merge_cells("A15:D15")
exceptions = [
    "전용 설비·유니언 사용 → 셋업비 30% 초과 예상 → 강제 고정형 분류",
    "포장기, 세척기, 포트 → 하이브리드 개조 불가 → 강제 고정형",
    "Mobile Shuttle 구조체 (MSS) → 유니언 사용 → 강제 고정형",
    "기존 설비 개조 case → 셋업비 비중 비정상적 高 → 산정 제외",
]
for i, ex in enumerate(exceptions):
    set_cell(ws, 16 + i, 1, f"E{i+1}", font=FONT_H2, fill=FILL_HEADER, align=ALIGN_C)
    ws.merge_cells(start_row=16+i, start_column=2, end_row=16+i, end_column=4)
    set_cell(ws, 16 + i, 2, ex, font=FONT_BODY, align=ALIGN_L)

set_cell(ws, 21, 1, "출처: F1 [39~44행]", font=FONT_FOOT, align=ALIGN_L)


# ============================================================
# Sheet 02 — LGES 사실
# ============================================================
ws = wb.create_sheet("02_LGES_사실")
for col, w in [("A", 24), ("B", 12), ("C", 12), ("D", 12), ("E", 12), ("F", 24)]:
    ws.column_dimensions[col].width = w

add_title_band(ws, 1, "LGES F1 공식 수치 (사내 사실, 변경 불가)", 6)

set_cell(ws, 3, 1, "★ 본 시트의 모든 수치는 LGES 공식 사내 자료(F1) — 변경 금지", font=FONT_NOTE, fill=FILL_KEY)
ws.merge_cells("A3:F3")

set_cell(ws, 5, 1, "사이트 / 연도", font=FONT_H2, fill=FILL_HEADER, align=ALIGN_C)
for j, y in enumerate(["'25년", "'26년", "'27년", "'28년"]):
    set_cell(ws, 5, j + 2, y, font=FONT_H2, fill=FILL_HEADER, align=ALIGN_C)
set_cell(ws, 5, 6, "출처", font=FONT_H2, fill=FILL_HEADER, align=ALIGN_C)

lges_rows = [
    ("LGES MI Lansing 유연성 지표", 0.47, 0.61, 0.75, 0.76, "F1 [48~50행]"),
    ("LGES ESGM2 각형 유연성 지표", 0.50, 0.62, 0.74, 0.78, "F1 [69~73행]"),
]
for i, r in enumerate(lges_rows):
    set_cell(ws, 6 + i, 1, r[0], font=FONT_BODY, fill=FILL_LGES, align=ALIGN_L)
    for j in range(4):
        set_cell(ws, 6 + i, j + 2, r[j + 1], font=FONT_INPUT, fill=FILL_LGES,
                 align=ALIGN_C, fmt="0.0%")
    set_cell(ws, 6 + i, 6, r[5], font=FONT_LINK, align=ALIGN_L)

set_cell(ws, 8, 1, "MI Lansing 추세 ('25→'28)", font=FONT_H2, fill=FILL_HEADER)
set_cell(ws, 8, 2, "=E6-B6", font=FONT_CALC, fill=FILL_LGES, align=ALIGN_C, fmt="0.0%")
ws.merge_cells("C8:E8")
set_cell(ws, 8, 3, "+29%p (47% → 76%) — 글로벌 각형 라인 중 명시적 유연성 지표 공개 유일", font=FONT_BODY, align=ALIGN_L)
set_cell(ws, 8, 6, "F1 [54행]", font=FONT_LINK, align=ALIGN_L)

set_cell(ws, 10, 1, "연도별 핵심 견인 과제", font=FONT_H1, fill=FILL_HEADER)
ws.merge_cells("A10:F10")
tasks = [
    ("'25년", "혁신물류 (SMA, Hybrid Stacker Crane, AMR)", "MI Lansing 1차 22.1% 절감 (329.3억)"),
    ("'26년", "전극 SMC (Smart Modular Cube)", "면적 25%↑, 설치공수 33%↓, 투자비 15%↓"),
    ("'27년", "MMF (Mobile Micro Factory, 활성화 SMC)", "로봇 기반 설비 조립 자동화"),
    ("'28년", "밀폐형 AMR (Cube 물류 설비)", "DR Less 공정 + 밀폐형 Cube Rack"),
]
set_cell(ws, 11, 1, "연도", font=FONT_H2, fill=FILL_HEADER, align=ALIGN_C)
ws.merge_cells("B11:C11")
set_cell(ws, 11, 2, "과제", font=FONT_H2, fill=FILL_HEADER, align=ALIGN_C)
ws.merge_cells("D11:E11")
set_cell(ws, 11, 4, "효과", font=FONT_H2, fill=FILL_HEADER, align=ALIGN_C)
set_cell(ws, 11, 6, "출처", font=FONT_H2, fill=FILL_HEADER, align=ALIGN_C)
for i, t in enumerate(tasks):
    set_cell(ws, 12 + i, 1, t[0], font=FONT_BODY, fill=FILL_LGES, align=ALIGN_C)
    ws.merge_cells(start_row=12+i, start_column=2, end_row=12+i, end_column=3)
    set_cell(ws, 12 + i, 2, t[1], font=FONT_BODY, fill=FILL_LGES, align=ALIGN_L)
    ws.merge_cells(start_row=12+i, start_column=4, end_row=12+i, end_column=5)
    set_cell(ws, 12 + i, 4, t[2], font=FONT_BODY, fill=FILL_LGES, align=ALIGN_L)
    set_cell(ws, 12 + i, 6, "F1 [69~73], F3, F4", font=FONT_LINK, align=ALIGN_L)

set_cell(ws, 17, 1, "MI Lansing 투자비 Breakdown (단위: 억원)", font=FONT_H1, fill=FILL_HEADER)
ws.merge_cells("A17:F17")
bd_headers = ["공정", "설비비", "셋업비", "합계", "셋업비 비중", "출처"]
for j, h in enumerate(bd_headers):
    set_cell(ws, 18, j + 1, h, font=FONT_H2, fill=FILL_HEADER, align=ALIGN_C)
bd_rows = [("전극", 192.5, 82.8), ("조립", 471.9, 139.1), ("활성화", 351.3, 84.7)]
for i, r in enumerate(bd_rows):
    set_cell(ws, 19 + i, 1, r[0], font=FONT_BODY, fill=FILL_LGES, align=ALIGN_C)
    set_cell(ws, 19 + i, 2, r[1], font=FONT_INPUT, fill=FILL_LGES, align=ALIGN_R, fmt="#,##0.0")
    set_cell(ws, 19 + i, 3, r[2], font=FONT_INPUT, fill=FILL_LGES, align=ALIGN_R, fmt="#,##0.0")
    set_cell(ws, 19 + i, 4, f"=B{19+i}+C{19+i}", font=FONT_CALC, fill=FILL_LGES, align=ALIGN_R, fmt="#,##0.0")
    set_cell(ws, 19 + i, 5, f"=C{19+i}/D{19+i}", font=FONT_CALC, fill=FILL_LGES, align=ALIGN_R, fmt="0.0%")
    set_cell(ws, 19 + i, 6, "F2 [76~80행]", font=FONT_LINK, align=ALIGN_L)
set_cell(ws, 22, 1, "Total", font=FONT_H2, fill=FILL_HEADER, align=ALIGN_C)
set_cell(ws, 22, 2, "=SUM(B19:B21)", font=FONT_CALC, fill=FILL_HEADER, align=ALIGN_R, fmt="#,##0.0")
set_cell(ws, 22, 3, "=SUM(C19:C21)", font=FONT_CALC, fill=FILL_HEADER, align=ALIGN_R, fmt="#,##0.0")
set_cell(ws, 22, 4, "=SUM(D19:D21)", font=FONT_CALC, fill=FILL_HEADER, align=ALIGN_R, fmt="#,##0.0")
set_cell(ws, 22, 5, "=C22/D22", font=FONT_CALC, fill=FILL_HEADER, align=ALIGN_R, fmt="0.0%")
set_cell(ws, 22, 6, "F2 [76~80행]", font=FONT_LINK, align=ALIGN_L)


# ============================================================
# 공통 함수 — 시트 03 / 04 (SDI / CATL) 빌더
# ============================================================
def build_competitor_sheet(ws, *, side, factories, weights, conservative_factor, optimistic_factor, sources, base_anchor):
    """
    side: 'SDI' or 'CATL'
    factories: list of (이름, 캐파, 폼팩터, 자동화 특성, [25,26,27,28] 유연성)
    weights: list of [25,26,27,28] GWh 가중치 (factories 와 동일 순서)
    conservative_factor / optimistic_factor: list of 4 (연도별 계수)
    """
    fill_main = FILL_SDI if side == "SDI" else FILL_CATL
    for col, w in [("A", 32), ("B", 12), ("C", 12), ("D", 12), ("E", 12), ("F", 22), ("G", 28)]:
        ws.column_dimensions[col].width = w

    add_title_band(ws, 1, f"{side} 유연성 추정 — 7공장 SUMPRODUCT 계산 + 시나리오 (AI 분석)", 7)

    set_cell(ws, 3, 1, f"★ {side} Base 시나리오 = SUMPRODUCT(공장별 자체 유연성, 공장별 GWh 가중치). 신뢰구간 ±5~10%p. 시나리오 변경 시 파란 입력값 수정.", font=FONT_NOTE, fill=FILL_KEY)
    ws.merge_cells("A3:G3")

    # § 1. 4-Proxy 가중치 (참고용)
    set_cell(ws, 5, 1, "§1. 4-Proxy 가중치 (가정 근거, 참고용)", font=FONT_H1, fill=FILL_HEADER)
    ws.merge_cells("A5:G5")

    proxy_h = ["Proxy", "가중치", "데이터 출처"]
    set_cell(ws, 6, 1, proxy_h[0], font=FONT_H2, fill=FILL_HEADER, align=ALIGN_C)
    ws.merge_cells("B6:C6")
    set_cell(ws, 6, 2, proxy_h[1], font=FONT_H2, fill=FILL_HEADER, align=ALIGN_C)
    ws.merge_cells("D6:G6")
    set_cell(ws, 6, 4, proxy_h[2], font=FONT_H2, fill=FILL_HEADER, align=ALIGN_C)

    proxies = [
        ("P1. AMR/AGV/Shuttle/OHT 도입 비중", 0.35, "설비 협력사 공시 (DART, 上交所), E3·E4·E6"),
        ("P2. Stacker Crane/Conveyor/MSS 비중", 0.30, "공장 설계 보도, WEF Lighthouse 자료"),
        ("P3. 신규 라인 vs 기존 개조 라인 비율", 0.20, "공장별 CAPA 증설 시점"),
        ("P4. 셋업 변경 시간 (Changeover Time)", 0.15, "WEF Lighthouse, 사례 분석"),
    ]
    for i, p in enumerate(proxies):
        set_cell(ws, 7 + i, 1, p[0], font=FONT_BODY, fill=fill_main, align=ALIGN_L)
        ws.merge_cells(start_row=7+i, start_column=2, end_row=7+i, end_column=3)
        set_cell(ws, 7 + i, 2, p[1], font=FONT_INPUT, fill=fill_main, align=ALIGN_C, fmt="0%")
        ws.merge_cells(start_row=7+i, start_column=4, end_row=7+i, end_column=7)
        set_cell(ws, 7 + i, 4, p[2], font=FONT_LINK, align=ALIGN_L)

    set_cell(ws, 11, 1, "합계 (검증)", font=FONT_H2, fill=FILL_HEADER)
    ws.merge_cells("B11:C11")
    set_cell(ws, 11, 2, "=SUM(B7:B10)", font=FONT_CALC, fill=FILL_HEADER, align=ALIGN_C, fmt="0%")

    # § 2. 7공장 × 4개년 자체 유연성 (입력값)
    set_cell(ws, 13, 1, "§2. 7공장 자체 유연성 ('25~'28, 입력값 — 시나리오 변경 시 조정)", font=FONT_H1, fill=FILL_HEADER)
    ws.merge_cells("A13:G13")

    f_headers = ["공장", "캐파/특성", "'25", "'26", "'27", "'28", "비고"]
    for j, h in enumerate(f_headers):
        set_cell(ws, 14, j + 1, h, font=FONT_H2, fill=FILL_HEADER, align=ALIGN_C)

    for i, f in enumerate(factories):
        name, capa, fmt_str, auto_char, y_flex = f
        set_cell(ws, 15 + i, 1, name, font=FONT_BODY, fill=fill_main, align=ALIGN_L)
        set_cell(ws, 15 + i, 2, f"{capa} / {fmt_str}", font=FONT_NOTE, fill=fill_main, align=ALIGN_L)
        for k in range(4):
            set_cell(ws, 15 + i, 3 + k, y_flex[k], font=FONT_INPUT, fill=fill_main,
                     align=ALIGN_C, fmt="0.0%")
        set_cell(ws, 15 + i, 7, auto_char, font=FONT_NOTE, fill=fill_main, align=ALIGN_L)

    # § 3. 7공장 × 4개년 GWh 가중치 (입력값)
    set_cell(ws, 23, 1, "§3. 7공장 GWh 가중치 ('25~'28, 입력값 — 합계 100%)", font=FONT_H1, fill=FILL_HEADER)
    ws.merge_cells("A23:G23")

    w_headers = ["공장", "—", "'25", "'26", "'27", "'28", "비고"]
    for j, h in enumerate(w_headers):
        set_cell(ws, 24, j + 1, h, font=FONT_H2, fill=FILL_HEADER, align=ALIGN_C)

    for i, w_row in enumerate(weights):
        name = factories[i][0]
        set_cell(ws, 25 + i, 1, name, font=FONT_BODY, fill=fill_main, align=ALIGN_L)
        set_cell(ws, 25 + i, 2, "—", font=FONT_NOTE, fill=fill_main, align=ALIGN_C)
        for k in range(4):
            set_cell(ws, 25 + i, 3 + k, w_row[k], font=FONT_INPUT, fill=fill_main,
                     align=ALIGN_C, fmt="0%")
        set_cell(ws, 25 + i, 7, "신규 라인 가동 시점 반영", font=FONT_NOTE, fill=fill_main, align=ALIGN_L)

    # 합계 검증 행
    set_cell(ws, 32, 1, "합계 (= 100% 확인)", font=FONT_H2, fill=FILL_HEADER, align=ALIGN_L)
    set_cell(ws, 32, 2, "—", font=FONT_NOTE, fill=FILL_HEADER, align=ALIGN_C)
    for k in range(4):
        col = get_column_letter(3 + k)
        set_cell(ws, 32, 3 + k, f"=SUM({col}25:{col}31)", font=FONT_CALC, fill=FILL_HEADER,
                 align=ALIGN_C, fmt="0%")

    # § 4. SUMPRODUCT 시나리오 결과
    set_cell(ws, 34, 1, f"§4. {side} 시나리오 결과 (SUMPRODUCT 자동 계산)", font=FONT_H1, fill=FILL_HEADER)
    ws.merge_cells("A34:G34")

    s_headers = ["시나리오", "수식 (참고)", "'25", "'26", "'27", "'28", "비고"]
    for j, h in enumerate(s_headers):
        set_cell(ws, 35, j + 1, h, font=FONT_H2, fill=FILL_HEADER, align=ALIGN_C)

    # Base 행 (36): SUMPRODUCT
    set_cell(ws, 36, 1, "Base (기준, 채택)", font=FONT_H2, fill=FILL_BASE, align=ALIGN_L)
    set_cell(ws, 36, 2, "=SUMPRODUCT(유연성, 가중치)", font=FONT_FORMULA, fill=FILL_BASE, align=ALIGN_L)
    for k in range(4):
        col = get_column_letter(3 + k)
        formula = f"=SUMPRODUCT({col}15:{col}21,{col}25:{col}31)"
        set_cell(ws, 36, 3 + k, formula, font=FONT_CALC, fill=FILL_BASE,
                 align=ALIGN_C, fmt="0.0%")
    set_cell(ws, 36, 7, "v2 본문 추정값 대비 ±2%p 오차 가능", font=FONT_NOTE, fill=FILL_BASE, align=ALIGN_L)

    # 보수 행 (37): Base × 보수 계수
    set_cell(ws, 37, 1, "보수 (Base × 보수 계수)", font=FONT_BODY, fill=fill_main, align=ALIGN_L)
    set_cell(ws, 37, 2, "=Base × C38:F38", font=FONT_FORMULA, fill=fill_main, align=ALIGN_L)
    for k in range(4):
        col = get_column_letter(3 + k)
        set_cell(ws, 37, 3 + k, f"={col}36*{col}38", font=FONT_CALC, fill=fill_main,
                 align=ALIGN_C, fmt="0.0%")
    set_cell(ws, 37, 7, "고정 라인 비중 高, 기존 개조 잔존", font=FONT_NOTE, fill=fill_main, align=ALIGN_L)

    # 보수 계수 행 (38, 입력값)
    set_cell(ws, 38, 1, "└ 보수 계수 (입력)", font=FONT_NOTE, fill=fill_main, align=ALIGN_L)
    set_cell(ws, 38, 2, "예: 0.88 ~ 0.91", font=FONT_NOTE, fill=fill_main, align=ALIGN_L)
    for k in range(4):
        set_cell(ws, 38, 3 + k, conservative_factor[k], font=FONT_INPUT, fill=fill_main,
                 align=ALIGN_C, fmt="0.00")

    # 낙관 행 (39): Base × 낙관 계수
    set_cell(ws, 39, 1, "낙관 (Base × 낙관 계수)", font=FONT_BODY, fill=fill_main, align=ALIGN_L)
    set_cell(ws, 39, 2, "=Base × C40:F40", font=FONT_FORMULA, fill=fill_main, align=ALIGN_L)
    for k in range(4):
        col = get_column_letter(3 + k)
        set_cell(ws, 39, 3 + k, f"={col}36*{col}40", font=FONT_CALC, fill=fill_main,
                 align=ALIGN_C, fmt="0.0%")
    set_cell(ws, 39, 7, "신규 라인 가속, 휴머노이드/AMR 확산", font=FONT_NOTE, fill=fill_main, align=ALIGN_L)

    # 낙관 계수 행 (40, 입력값)
    set_cell(ws, 40, 1, "└ 낙관 계수 (입력)", font=FONT_NOTE, fill=fill_main, align=ALIGN_L)
    set_cell(ws, 40, 2, "예: 1.09 ~ 1.14", font=FONT_NOTE, fill=fill_main, align=ALIGN_L)
    for k in range(4):
        set_cell(ws, 40, 3 + k, optimistic_factor[k], font=FONT_INPUT, fill=fill_main,
                 align=ALIGN_C, fmt="0.00")

    # v2 본문 참조값 (검증용)
    set_cell(ws, 42, 1, "참고: v2 본문 추정값 (정합성 검증용)", font=FONT_NOTE, fill=FILL_HEADER, align=ALIGN_L)
    set_cell(ws, 42, 2, "보수/Base/낙관", font=FONT_NOTE, fill=FILL_HEADER, align=ALIGN_C)
    for k in range(4):
        b, m, o = base_anchor[k]
        set_cell(ws, 42, 3 + k, f"{int(b*100)}/{int(m*100)}/{int(o*100)}%",
                 font=FONT_NOTE, fill=FILL_HEADER, align=ALIGN_C)
    set_cell(ws, 42, 7, "v2 .md 본문 §4.4 / §5.5", font=FONT_LINK, align=ALIGN_L)

    # 안내
    set_cell(ws, 44, 1, "★ 시트 '05_12셀_매트릭스' 의 " + side + " 행은 본 시트 36행 (Base) 을 자동 참조", font=FONT_NOTE, fill=FILL_KEY)
    ws.merge_cells("A44:G44")

    # 출처
    set_cell(ws, 46, 1, "출처", font=FONT_H1, fill=FILL_HEADER)
    ws.merge_cells("A46:G46")
    for i, s in enumerate(sources):
        ws.merge_cells(start_row=47+i, start_column=1, end_row=47+i, end_column=7)
        set_cell(ws, 47 + i, 1, s, font=FONT_LINK, align=ALIGN_L)


# ============================================================
# Sheet 03 — SDI
# ============================================================
ws = wb.create_sheet("03_SDI_계산")

sdi_factories = [
    # (이름, 캐파, 폼팩터, 자동화 특성, ['25,'26,'27,'28] 유연성)
    ("천안 (46파이 마더라인)", "1~2 GWh", "46파이 원통형", "신규, 국산 장비 위주", [0.46, 0.52, 0.62, 0.70]),
    ("울산 (각형)", "~30 GWh", "P5/P6 각형", "기존 Conveyor 중심 일부 P6 전환", [0.31, 0.33, 0.36, 0.40]),
    ("헝가리 괴드 1·2", "40 GWh", "P5/P6 각형", "P5→P6 전환 60→95%", [0.33, 0.45, 0.54, 0.62]),
    ("미국 코코모 1 (SPE)", "33 GWh", "P6 각형 NCA+LFP", "신규 라인, IRA 대응", [0.41, 0.45, 0.51, 0.58]),
    ("미국 코코모 2 (SPE, '27)", "34 GWh", "P6 각형", "신규 라인", [0.00, 0.00, 0.485, 0.55]),
    ("미국 뉴칼라일 (GM JV, '27)", "27→36 GWh", "각형 + 원통형 혼재", "신규, 다양한 폼팩터", [0.00, 0.00, 0.515, 0.60]),
    ("말레이시아 SDIEM", "—", "21700 원통형", "기존 + 신규 BBU", [0.375, 0.40, 0.43, 0.46]),
]
# 가중치 = 신규 라인 가동 곡선 반영 (합계 100%)
sdi_weights = [
    [0.07, 0.07, 0.10, 0.14],   # 천안
    [0.30, 0.30, 0.24, 0.18],   # 울산
    [0.42, 0.40, 0.32, 0.24],   # 괴드
    [0.18, 0.17, 0.15, 0.14],   # 코코모 1
    [0.00, 0.00, 0.07, 0.13],   # 코코모 2
    [0.00, 0.03, 0.09, 0.14],   # 뉴칼라일
    [0.03, 0.03, 0.03, 0.03],   # SDIEM
]
sdi_conservative = [0.91, 0.88, 0.86, 0.88]   # 연도별 보수 계수
sdi_optimistic = [1.09, 1.07, 1.06, 1.09]     # 연도별 낙관 계수
sdi_anchor = [(0.32, 0.35, 0.38), (0.36, 0.41, 0.44), (0.42, 0.49, 0.52), (0.50, 0.57, 0.62)]
sdi_sources = [
    "D4: references/경쟁사/2026-05-21_SamsungSDI_StarPlus_DOE_LPO.md",
    "D5: references/경쟁사/2026-05-21_SamsungSDI_P6_Hungary_InterBattery2025.md",
    "E1: electrive.com 2025.5.27 SDI Hungary 각형 전환",
    "E3: thelec.kr SFA 1Q 수주 2,179억 (+18.3% YoY)",
    "E4: hellot.net 코윈테크 AMR 수주 +465.2% YoY (108302)",
    "E6: m.joseilbo.com 씨케이솔루션 SDI 헝가리 €16.8m 수주",
    "F5: references/경쟁사/LGES_FA_Prismatic_Benchmarking_Integrated_v1-v5.md [1077~1087행]",
]
build_competitor_sheet(ws, side="SDI", factories=sdi_factories, weights=sdi_weights,
                       conservative_factor=sdi_conservative,
                       optimistic_factor=sdi_optimistic,
                       sources=sdi_sources, base_anchor=sdi_anchor)


# ============================================================
# Sheet 04 — CATL
# ============================================================
ws = wb.create_sheet("04_CATL_계산")

catl_factories = [
    ("푸젠 닝더 본사", "다수 GWh", "각형 다종", "PSL 8세대, 고정형 Conveyor + Stacker", [0.21, 0.24, 0.28, 0.34]),
    ("푸젠 푸딩 No.5 (다층)", "25 GWh", "각형", "다층, Stacker Crane 의존 高", [0.175, 0.20, 0.24, 0.30]),
    ("쓰촨 이빈 (CATL-SC)", "16 GWh", "각형", "Pack 80% 자동화, gluing 인력 -70%", [0.275, 0.31, 0.37, 0.44]),
    ("장쑤 리양 (4단계)", "126 GWh", "각형", "Additive Mfg, 출력 +320%", [0.31, 0.35, 0.41, 0.48]),
    ("독일 에르푸르트 (CATT)", "8→14→24 GWh", "각형 + 모듈", "6 라인, 1,700명", [0.25, 0.28, 0.33, 0.40]),
    ("헝가리 데브레첸 Phase 1", "40 GWh ('26 Q1 SOP)", "각형", "BMW/MB/Stellantis/VW 공급", [0.00, 0.285, 0.40, 0.50]),
    ("허난 뤄양 (Pack)", "—", "Pack 전용", "Xiaomo 휴머노이드 EOL/DCR", [0.23, 0.27, 0.35, 0.45]),
]
catl_weights = [
    [0.32, 0.28, 0.24, 0.20],   # 닝더
    [0.13, 0.13, 0.12, 0.11],   # 푸딩
    [0.09, 0.08, 0.08, 0.07],   # 이빈
    [0.30, 0.30, 0.28, 0.26],   # 리양
    [0.06, 0.07, 0.08, 0.08],   # 에르푸르트
    [0.00, 0.06, 0.12, 0.20],   # 데브레첸
    [0.10, 0.08, 0.08, 0.08],   # 뤄양
]
catl_conservative = [0.88, 0.83, 0.80, 0.76]
catl_optimistic = [1.12, 1.14, 1.14, 1.10]
catl_anchor = [(0.22, 0.25, 0.28), (0.24, 0.29, 0.33), (0.28, 0.35, 0.40), (0.32, 0.42, 0.46)]
catl_sources = [
    "D1: references/경쟁사/2026-05-21_CATL_Debrecen_LEAD_GWh환산.md",
    "D2: references/경쟁사/2026-05-21_CATL_PSL_8세대_SmartManufacturing.md (인력 -70%, 속도 +300%)",
    "D3: references/경쟁사/2026-05-21_CATL_Lighthouse_Sichuan_Liyang.md",
    "D8: catl.com/en/news/6030.html — 8세대 PSL 공식",
    "D9: catl.com/en/manufacture — 95% 인터커넥션",
    "E11: cleantechnica.com 2025.12 — Xiaomo 휴머노이드",
    "E12: cninfo.com.cn — 先导智能 2024 연간보고서 (智能物流 23.8%)",
    "F5: references/경쟁사/LGES_FA_Prismatic_Benchmarking_Integrated_v1-v5.md [1053~1058, 1060~1072행]",
]
build_competitor_sheet(ws, side="CATL", factories=catl_factories, weights=catl_weights,
                       conservative_factor=catl_conservative,
                       optimistic_factor=catl_optimistic,
                       sources=catl_sources, base_anchor=catl_anchor)


# ============================================================
# Sheet 05 — 12셀 매트릭스 (시트 02·03·04 참조)
# ============================================================
ws = wb.create_sheet("05_12셀_매트릭스")
for col, w in [("A", 30), ("B", 12), ("C", 12), ("D", 12), ("E", 12), ("F", 16), ("G", 14)]:
    ws.column_dimensions[col].width = w

add_title_band(ws, 1, "3사 × 4개년 통합 매트릭스 (12셀) — 자동 갱신", 7)

set_cell(ws, 3, 1, "★ LGES 행 = '02_LGES_사실' 참조 / SDI 행 = '03_SDI_계산'!B36:E36 (Base SUMPRODUCT) / CATL 행 = '04_CATL_계산'!B36:E36", font=FONT_NOTE, fill=FILL_KEY)
ws.merge_cells("A3:G3")

set_cell(ws, 5, 1, "사 / 사이트", font=FONT_H2, fill=FILL_HEADER, align=ALIGN_C)
for j, y in enumerate(["'25년", "'26년", "'27년", "'28년"]):
    set_cell(ws, 5, j + 2, y, font=FONT_H2, fill=FILL_HEADER, align=ALIGN_C)
set_cell(ws, 5, 6, "신뢰구간", font=FONT_H2, fill=FILL_HEADER, align=ALIGN_C)
set_cell(ws, 5, 7, "성격", font=FONT_H2, fill=FILL_HEADER, align=ALIGN_C)

# LGES MI Lansing
set_cell(ws, 6, 1, "LGES MI Lansing (F1 공식)", font=FONT_BODY, fill=FILL_LGES, align=ALIGN_L)
for j in range(4):
    col_letter = get_column_letter(j + 2)
    set_cell(ws, 6, j + 2, f"='02_LGES_사실'!{col_letter}6",
             font=FONT_CALC, fill=FILL_LGES, align=ALIGN_C, fmt="0.0%")
set_cell(ws, 6, 6, "±0%", font=FONT_BODY, fill=FILL_LGES, align=ALIGN_C)
set_cell(ws, 6, 7, "사내 사실", font=FONT_H2, fill=FILL_LGES, align=ALIGN_C)

# LGES ESGM2
set_cell(ws, 7, 1, "LGES ESGM2 각형 (F1 공식)", font=FONT_BODY, fill=FILL_LGES, align=ALIGN_L)
for j in range(4):
    col_letter = get_column_letter(j + 2)
    set_cell(ws, 7, j + 2, f"='02_LGES_사실'!{col_letter}7",
             font=FONT_CALC, fill=FILL_LGES, align=ALIGN_C, fmt="0.0%")
set_cell(ws, 7, 6, "±0%", font=FONT_BODY, fill=FILL_LGES, align=ALIGN_C)
set_cell(ws, 7, 7, "사내 사실", font=FONT_H2, fill=FILL_LGES, align=ALIGN_C)

# SDI Base (시트 03 행 36 = SUMPRODUCT)
set_cell(ws, 8, 1, "Samsung SDI (Base, SUMPRODUCT)", font=FONT_BODY, fill=FILL_SDI, align=ALIGN_L)
for j in range(4):
    col_letter = get_column_letter(j + 2)
    set_cell(ws, 8, j + 2, f"='03_SDI_계산'!{col_letter}36",
             font=FONT_CALC, fill=FILL_SDI, align=ALIGN_C, fmt="0.0%")
set_cell(ws, 8, 6, "±5~8%p", font=FONT_BODY, fill=FILL_SDI, align=ALIGN_C)
set_cell(ws, 8, 7, "AI 추정", font=FONT_H2, fill=FILL_SDI, align=ALIGN_C)

# CATL Base
set_cell(ws, 9, 1, "CATL (Base, SUMPRODUCT)", font=FONT_BODY, fill=FILL_CATL, align=ALIGN_L)
for j in range(4):
    col_letter = get_column_letter(j + 2)
    set_cell(ws, 9, j + 2, f"='04_CATL_계산'!{col_letter}36",
             font=FONT_CALC, fill=FILL_CATL, align=ALIGN_C, fmt="0.0%")
set_cell(ws, 9, 6, "±5~10%p", font=FONT_BODY, fill=FILL_CATL, align=ALIGN_C)
set_cell(ws, 9, 7, "AI 추정", font=FONT_H2, fill=FILL_CATL, align=ALIGN_C)

# 격차 행 (자동)
set_cell(ws, 11, 1, "LGES vs SDI 격차 (MI Lansing 기준)", font=FONT_H2, fill=FILL_HEADER, align=ALIGN_L)
for j in range(4):
    col_letter = get_column_letter(j + 2)
    set_cell(ws, 11, j + 2, f"={col_letter}6-{col_letter}8",
             font=FONT_CALC, fill=FILL_HEADER, align=ALIGN_C, fmt="0.0%")
set_cell(ws, 11, 6, "—", font=FONT_BODY, align=ALIGN_C)
set_cell(ws, 11, 7, "수식 계산", font=FONT_H2, align=ALIGN_C)

set_cell(ws, 12, 1, "LGES vs CATL 격차 (MI Lansing 기준)", font=FONT_H2, fill=FILL_HEADER, align=ALIGN_L)
for j in range(4):
    col_letter = get_column_letter(j + 2)
    set_cell(ws, 12, j + 2, f"={col_letter}6-{col_letter}9",
             font=FONT_CALC, fill=FILL_HEADER, align=ALIGN_C, fmt="0.0%")
set_cell(ws, 12, 6, "—", font=FONT_BODY, align=ALIGN_C)
set_cell(ws, 12, 7, "수식 계산", font=FONT_H2, align=ALIGN_C)

# 사용 가이드
set_cell(ws, 15, 1, "사용 가이드 — 시나리오 변경 방법", font=FONT_H1, fill=FILL_HEADER)
ws.merge_cells("A15:G15")
guides = [
    "1. SDI Base 변경: '03_SDI_계산' §2 (15~21행 자체 유연성) 또는 §3 (25~31행 가중치) 의 파란 입력값 수정 → 본 시트 SDI 행 자동 갱신",
    "2. CATL Base 변경: '04_CATL_계산' 동일 방식",
    "3. 보수/낙관 조정: 시트 03·04 의 38·40행 (보수·낙관 계수, 파란 입력) 수정",
    "4. LGES 행은 F1 사내 사실 — 변경 금지",
    "5. 격차는 LGES MI Lansing 기준. ESGM2 기준 격차는 ='02_LGES_사실'!B7 - SDI/CATL 행 수식 추가",
]
for i, g in enumerate(guides):
    ws.merge_cells(start_row=16+i, start_column=1, end_row=16+i, end_column=7)
    set_cell(ws, 16 + i, 1, g, font=FONT_BODY, align=ALIGN_L)


# ============================================================
# Sheet 06 — 가정 카드
# ============================================================
ws = wb.create_sheet("06_가정카드")
for col, w in [("A", 8), ("B", 52), ("C", 22), ("D", 32)]:
    ws.column_dimensions[col].width = w

add_title_band(ws, 1, "가정 카드 A1~A10 (SDI/CATL 추정의 근거 가정)", 4)

set_cell(ws, 3, 1, "★ 본 가정 변경 시 시트 03·04 의 입력값(파랑)을 수동 조정 필요. 본 시트는 추적용.", font=FONT_NOTE, fill=FILL_KEY)
ws.merge_cells("A3:D3")

headers = ["ID", "가정 내용", "영향 셀", "근거 출처"]
for j, h in enumerate(headers):
    set_cell(ws, 5, j + 1, h, font=FONT_H2, fill=FILL_HEADER, align=ALIGN_C)

assumptions = [
    ("A1", "CATL Debrecen €7.3B 중 EPC·건축비 30% 차감 → 자동화 단가 환산", "04 데브레첸 행 (20)", "D1 [42~43행]"),
    ("A2", "LEAD Intelligent 모듈 95% / Pack 50% 자동화 표준이 Debrecen 적용", "04 데브레첸 행 (20)", "D1 [45행]"),
    ("A3", "CATL 8세대 PSL 인력 -70% → AMR/OHT 비중 약 +20%p", "04 닝더·이빈 행", "D2 [29, 44행]"),
    ("A4", "SDI StarPlus 코코모 = 한국 EPC + 동일 공급사 → LGES ±5%p 밴드", "03 코코모 1·2 행", "F5 [1077~1087]"),
    ("A5", '"CATL 95% 인터커넥션" ≠ "유연 설비 95%" (IoT 연결률)', "04 전체", "D2 [50~57]"),
    ("A6", "CATL Yibin 팩 80%는 팩 한정 (셀/모듈 아님) → Debrecen 환산 제외", "04 이빈 행", "D3 [29~31, 42~46]"),
    ("A7", "SDI P5 wound→P6 stacked 전환 → 셋업비 30% 초과로 고정 분류", "03 괴드 1·2 행", "E1, E14"),
    ("A8", "천안 46파이 마더라인 → 신규 폼팩터, 셋업비 ≤30% 통과", "03 천안 행", "산업 매체 추정"),
    ("A9", "Xiaomo 휴머노이드 ('25.12 뤄양) = Pack 한정. 셀 라인 확대 미정", "04 뤄양 / 낙관 시나리오", "E11"),
    ("A10", "F1 vs F2 (70:30 vs 75:25) → F1 채택. 30% 컷오프와 정합", "01 정의·분류표", "F1 [35] vs F2 [87]"),
]
for i, a in enumerate(assumptions):
    set_cell(ws, 6 + i, 1, a[0], font=FONT_H2, fill=FILL_KEY, align=ALIGN_C)
    set_cell(ws, 6 + i, 2, a[1], font=FONT_BODY, align=ALIGN_L)
    set_cell(ws, 6 + i, 3, a[2], font=FONT_BODY, align=ALIGN_L)
    set_cell(ws, 6 + i, 4, a[3], font=FONT_LINK, align=ALIGN_L)


# ============================================================
# Sheet 07 — 공정별
# ============================================================
ws = wb.create_sheet("07_공정별_매트릭스")
for col, w in [("A", 22), ("B", 22), ("C", 22), ("D", 22), ("E", 28)]:
    ws.column_dimensions[col].width = w

add_title_band(ws, 1, "공정 5단계 × 3사 자동화 성숙도", 5)

set_cell(ws, 3, 1, "★ LGES = 유연 설비 우위 / SDI = 표준화 중간 / CATL = 고정형 자동화", font=FONT_NOTE, fill=FILL_KEY)
ws.merge_cells("A3:E3")

p_headers = ["공정", "LGES 유연성", "SDI 유연성", "CATL 유연성", "비고"]
for j, h in enumerate(p_headers):
    set_cell(ws, 5, j + 1, h, font=FONT_H2, fill=FILL_HEADER, align=ALIGN_C)

processes = [
    ("전극 (믹싱/코팅/압연/슬리팅)", "'26 SMC 도입 (高)", "한화모멘텀/피엔티 표준 (中)", "LEAD 涂布/辊压 고정형 (低)", "CATL 동일 화학 가정"),
    ("조립 (노칭/스태킹/조립)", "高 (MMF 적용)", "中 (필에너지/엠오티)", "低 (PSL 8세대 고속 라인)", "LGES 우위 명확"),
    ("화성/에이징", "'27 활성화 SMC (高)", "원익피앤이/갑진 (中)", "LEAD 化成容量 고정형 (低)", "LGES 차별화"),
    ("팩 (Pack)", "高 (모듈형)", "中", "Xiaomo 휴머노이드 (中, '25.12~)", "CATL 후공정 추격"),
    ("물류 (반제품/완제품)", "'28 밀폐형 AMR (高)", "SFA AGV + AS-RS (中)", "Geek+/Hai/Quicktron + Conveyor (低)", "LGES 혁신물류"),
]
for i, p in enumerate(processes):
    set_cell(ws, 6 + i, 1, p[0], font=FONT_H2, fill=FILL_HEADER, align=ALIGN_L)
    set_cell(ws, 6 + i, 2, p[1], font=FONT_BODY, fill=FILL_LGES, align=ALIGN_L)
    set_cell(ws, 6 + i, 3, p[2], font=FONT_BODY, fill=FILL_SDI, align=ALIGN_L)
    set_cell(ws, 6 + i, 4, p[3], font=FONT_BODY, fill=FILL_CATL, align=ALIGN_L)
    set_cell(ws, 6 + i, 5, p[4], font=FONT_NOTE, align=ALIGN_L)

set_cell(ws, 12, 1, "출처: outputs/유연성_경쟁사_분석_v2_2026-05-21.md §6.2 + F5 [1053~1058행]", font=FONT_FOOT, align=ALIGN_L)
ws.merge_cells("A12:E12")


# ============================================================
# Sheet 08 — 라인 재구성 비용
# ============================================================
ws = wb.create_sheet("08_라인재구성_비용")
for col, w in [("A", 32), ("B", 18), ("C", 18), ("D", 22)]:
    ws.column_dimensions[col].width = w

add_title_band(ws, 1, "라인 재구성 비용 비교 (3 시나리오) + 셋업 변경 시간", 4)

set_cell(ws, 3, 1, "★ 신규 폼팩터 도입 시 LGES 재구성 비용은 CATL 대비 약 1/3 수준 (SMC 재활용)", font=FONT_NOTE, fill=FILL_KEY)
ws.merge_cells("A3:D3")

c_headers = ["시나리오", "LGES", "SDI", "CATL"]
for j, h in enumerate(c_headers):
    set_cell(ws, 5, j + 1, h, font=FONT_H2, fill=FILL_HEADER, align=ALIGN_C)

cost_scenarios = [
    ("동일 폼팩터 화학 변경 (NMC→LFP)", "5~10%", "12~18%", "20~30% (라인 분리 신축)"),
    ("폼팩터 부분 변경 (셀 크기)", "8~15%", "18~25%", "25~35%"),
    ("신규 폼팩터 도입 (46파이)", "15~25% (SMC 재활용)", "25~35% (천안 신규)", "35~50% (전용 공장 신축)"),
]
for i, s in enumerate(cost_scenarios):
    set_cell(ws, 6 + i, 1, s[0], font=FONT_H2, fill=FILL_HEADER, align=ALIGN_L)
    set_cell(ws, 6 + i, 2, s[1], font=FONT_BODY, fill=FILL_LGES, align=ALIGN_C)
    set_cell(ws, 6 + i, 3, s[2], font=FONT_BODY, fill=FILL_SDI, align=ALIGN_C)
    set_cell(ws, 6 + i, 4, s[3], font=FONT_BODY, fill=FILL_CATL, align=ALIGN_C)

set_cell(ws, 10, 1, "셋업 변경 시간 (상대값)", font=FONT_H1, fill=FILL_HEADER)
ws.merge_cells("A10:D10")
for j, h in enumerate(["지표", "LGES", "SDI", "CATL"]):
    set_cell(ws, 11, j + 1, h, font=FONT_H2, fill=FILL_HEADER, align=ALIGN_C)
set_cell(ws, 12, 1, "셋업 변경 시간", font=FONT_H2, fill=FILL_HEADER, align=ALIGN_L)
set_cell(ws, 12, 2, "1.0x (기준)", font=FONT_BODY, fill=FILL_LGES, align=ALIGN_C)
set_cell(ws, 12, 3, "1.4x", font=FONT_BODY, fill=FILL_SDI, align=ALIGN_C)
set_cell(ws, 12, 4, "1.8x", font=FONT_BODY, fill=FILL_CATL, align=ALIGN_C)
set_cell(ws, 13, 1, "1초/셀 달성", font=FONT_H2, fill=FILL_HEADER, align=ALIGN_L)
set_cell(ws, 13, 2, "부분 (PPM)", font=FONT_BODY, fill=FILL_LGES, align=ALIGN_C)
set_cell(ws, 13, 3, "미달성", font=FONT_BODY, fill=FILL_SDI, align=ALIGN_C)
set_cell(ws, 13, 4, "달성 (PSL 8세대)", font=FONT_BODY, fill=FILL_CATL, align=ALIGN_C)

set_cell(ws, 15, 1, "출처: outputs/유연성_경쟁사_분석_v2_2026-05-21.md §6.3 + D2 [29~31행]", font=FONT_FOOT, align=ALIGN_L)
ws.merge_cells("A15:D15")


# ============================================================
# Sheet 09 — 출처
# ============================================================
ws = wb.create_sheet("09_출처")
for col, w in [("A", 10), ("B", 55), ("C", 18), ("D", 60)]:
    ws.column_dimensions[col].width = w

add_title_band(ws, 1, "출처 통합 — L1 사내 사실 / L2~L3 사내 분석 / L4 외부", 4)

set_cell(ws, 3, 1, "코드", font=FONT_H2, fill=FILL_HEADER, align=ALIGN_C)
set_cell(ws, 3, 2, "출처 명", font=FONT_H2, fill=FILL_HEADER, align=ALIGN_C)
set_cell(ws, 3, 3, "카테고리", font=FONT_H2, fill=FILL_HEADER, align=ALIGN_C)
set_cell(ws, 3, 4, "URL / 위치", font=FONT_H2, fill=FILL_HEADER, align=ALIGN_C)

sources = [
    ("F1", "FA 유연성 지표 정의 (LGES 공식, 2026-05-21, 74줄)", "L1 사내 사실 (유일)", "references/FA 유연성 지표 정의"),
    ("F2", "flexibility_indicator_summary.md (2026-05-14)", "L2 사내 분석", "references/용어집/flexibility_indicator_summary.md"),
    ("F3", "26FA KPI.md (Slide 1~11)", "L2 사내 분석", "references/26FA KPI.md"),
    ("F4", "2026 FA기술담당 중장기로드맵 v2", "L2 사내 분석", "references/roadmap/2026_FA기술담당_중장기로드맵_v2.md"),
    ("F5", "LGES FA Prismatic Benchmarking v1-v5 (1,276줄)", "L2 사내 분석", "references/경쟁사/LGES_FA_Prismatic_Benchmarking_Integrated_v1-v5.md"),
    ("D1", "CATL Debrecen LEAD GWh 환산 (2026-05-21)", "L3 외부 IR 아카이브", "references/경쟁사/2026-05-21_CATL_Debrecen_LEAD_GWh환산.md"),
    ("D2", "CATL PSL 8세대 Smart Manufacturing (2026-05-21)", "L3 외부 IR 아카이브", "references/경쟁사/2026-05-21_CATL_PSL_8세대_SmartManufacturing.md"),
    ("D3", "CATL Lighthouse Sichuan/Liyang (2026-05-21)", "L3 외부 IR 아카이브", "references/경쟁사/2026-05-21_CATL_Lighthouse_Sichuan_Liyang.md"),
    ("D4", "Samsung SDI StarPlus DOE LPO (2026-05-21)", "L3 외부 IR 아카이브", "references/경쟁사/2026-05-21_SamsungSDI_StarPlus_DOE_LPO.md"),
    ("D5", "Samsung SDI P6 Hungary InterBattery 2025", "L3 외부 IR 아카이브", "references/경쟁사/2026-05-21_SamsungSDI_P6_Hungary_InterBattery2025.md"),
    ("D6", "CATL Debrecen €7.3B 발표 (Fozlsg)", "L4 외부 공식", "https://en.fozlsg.com/News-DE-dt/32.html"),
    ("D7", "Debrecen SOP '26 Q1", "L4 외부 공식", "https://electrification-solutions.com/catl-hungary-plant-set-to-begin-production-in-early-2026/"),
    ("D8", "CATL 8세대 PSL 인력 -70%·속도 +300%", "L4 외부 공식", "https://www.catl.com/en/news/6030.html"),
    ("D9", "CATL Smart Manufacturing 95% 인터커넥션", "L4 외부 공식", "https://www.catl.com/en/manufacture/"),
    ("D10", "DOE LPO $7.54B StarPlus", "L4 외부 공식", "https://www.energy.gov/edf/articles/lpo-announces-754-billion-loan-starplus-energy-construct-lithium-ion-battery-factories"),
    ("E1", "SDI 헝가리 각형 전환 ('25.5.27)", "L4 외부 매체", "https://www.electrive.com/2025/05/27/samsung-sdi-paves-way-for-prismatic-batteries-at-hungary-plant/"),
    ("E2", "SDI Hungary 증설 (Battery-News, '25.6.3)", "L4 외부 매체", "https://battery-news.de/en/2025/06/"),
    ("E3", "에스에프에이 1Q 수주 2,179억 (+18.3% YoY)", "L4 외부 매체", "https://www.thelec.kr/news/articleView.html?idxno=7953"),
    ("E4", "코윈테크 AMR 수주 +465.2% YoY (헬로티 108302)", "L4 외부 매체", "https://www.hellot.net/news/article.html?no=108302"),
    ("E5", "코윈테크 320억 자동화 시스템 턴키 (헬로티 97091)", "L4 외부 매체", "https://www.hellot.net/news/article.html?no=97091"),
    ("E6", "씨케이솔루션 SDI 헝가리 €16.8m 수주", "L4 외부 매체", "https://m.joseilbo.com/news/view.htm?newsid=567743"),
    ("E7", "씨케이솔루션 SDI 헝가리 €13.93m 수주", "L4 외부 매체", "https://www.digitaltoday.co.kr/news/articleView.html?idxno=571068"),
    ("E8", "CATL Sichuan Lighthouse 인증", "L4 외부 매체", "https://www.catl.com/en/news/1030.html"),
    ("E9", "CATL Liyang Lighthouse 인증", "L4 외부 매체", "https://www.catl.com/en/news/6176.html"),
    ("E10", "CATL Yibin Sustainability Lighthouse", "L4 외부 매체", "https://www.catl.com/en/news/6652.html"),
    ("E11", "CATL Xiaomo 휴머노이드 ('25.12 뤄양)", "L4 외부 매체", "https://cleantechnica.com/2025/12/24/humanoid-robo"),
    ("E12", "先导智能 2024 연간보고서 (cninfo, 智能物流 23.8%)", "L4 외부 매체", "https://static.cninfo.com.cn/finalpage/2025-04-29/"),
    ("E13", "SDI 미국 LFP 라인 전환 (Solar Power World)", "L4 외부 매체", "https://www.solarpowerworldonline.com/"),
    ("E14", "SDI 헝가리 각형 전환 (The Elec)", "L4 외부 매체", "https://www.thelec.kr/news/articleView.html"),
    ("E15", "CATL Smart Manufacturing 공식", "L4 외부 매체", "https://www.catl.com/en/manufacture"),
]
for i, s in enumerate(sources):
    if s[0] == "F1":
        fill = FILL_KEY
    elif s[0].startswith("F"):
        fill = FILL_LGES
    elif s[0].startswith("D"):
        fill = FILL_CATL
    else:
        fill = FILL_SDI
    set_cell(ws, 4 + i, 1, s[0], font=FONT_H2, fill=fill, align=ALIGN_C)
    set_cell(ws, 4 + i, 2, s[1], font=FONT_BODY, align=ALIGN_L)
    set_cell(ws, 4 + i, 3, s[2], font=FONT_NOTE, align=ALIGN_C)
    set_cell(ws, 4 + i, 4, s[3], font=FONT_LINK, align=ALIGN_L)

foot_row = 4 + len(sources) + 2
set_cell(ws, foot_row, 1, "※ 본 백데이터는 outputs/유연성_경쟁사_분석_v2_2026-05-21.md (449줄) 의 정량 추정을 재구성한 것임. v2 = 최종 결정판.", font=FONT_FOOT, align=ALIGN_L)
ws.merge_cells(start_row=foot_row, start_column=1, end_row=foot_row, end_column=4)


# ============================================================
# 저장
# ============================================================
OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)

size_kb = OUT.stat().st_size / 1024
print(f"saved -> {OUT}")
print(f"size  -> {size_kb:.1f} KB")
print(f"sheets -> {len(wb.sheetnames)} ({', '.join(wb.sheetnames)})")

# 검증 — 수식 카운트
from openpyxl import load_workbook
wb2 = load_workbook(OUT)
total_formulas = 0
for sn in wb2.sheetnames:
    ws = wb2[sn]
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                total_formulas += 1
print(f"formulas -> {total_formulas} 개")
