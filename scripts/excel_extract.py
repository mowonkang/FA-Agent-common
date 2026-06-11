"""Excel(.xlsx) 파일을 markdown 으로 변환한다.

각 시트는 H2(`## 시트명`) 섹션으로, 셀 내용은 markdown 표로 출력된다.
1행은 헤더로 처리한다. 병합 셀은 좌상단 값을 모든 영역에 복제한다.
수식은 워크북에 저장된 캐시된 값을 사용한다(data_only=True).

사용법:
    python3 scripts/excel_extract.py <input.xlsx> [--out <output.md>] [--sheet <name>]

기본 출력 위치는 입력 파일과 같은 폴더의 `<basename>.md` 이다.
`--sheet` 로 특정 시트만 추출할 수 있다.
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.stderr.write(
        "openpyxl 이 설치되어 있지 않습니다. 먼저 다음을 실행하세요:\n"
        "    pip install -r requirements.txt\n"
    )
    sys.exit(2)


def _cell_to_text(value: Any) -> str:
    """셀 값을 markdown 표에 들어갈 문자열로 변환."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    text = str(value)
    return text.replace("\\", "\\\\").replace("|", "\\|").replace("\r\n", "\n").replace("\n", "<br>")


def _expand_merged_cells(worksheet) -> dict[tuple[int, int], Any]:
    """병합 셀의 좌상단 값을 모든 병합 영역 좌표에 복제한 dict 반환.

    반환: {(row, col): value} — 1-based 좌표.
    """
    overrides: dict[tuple[int, int], Any] = {}
    for merged_range in worksheet.merged_cells.ranges:
        top_left = worksheet.cell(row=merged_range.min_row, column=merged_range.min_col).value
        for r in range(merged_range.min_row, merged_range.max_row + 1):
            for c in range(merged_range.min_col, merged_range.max_col + 1):
                overrides[(r, c)] = top_left
    return overrides


def _sheet_to_rows(worksheet) -> list[list[str]]:
    """시트의 셀 값을 2D 리스트(문자열)로 반환. 병합 셀은 펼침."""
    overrides = _expand_merged_cells(worksheet)
    max_row = worksheet.max_row or 0
    max_col = worksheet.max_column or 0
    rows: list[list[str]] = []
    for r in range(1, max_row + 1):
        row: list[str] = []
        for c in range(1, max_col + 1):
            if (r, c) in overrides:
                value = overrides[(r, c)]
            else:
                value = worksheet.cell(row=r, column=c).value
            row.append(_cell_to_text(value))
        rows.append(row)
    return rows


def _trim_empty_edges(rows: list[list[str]]) -> list[list[str]]:
    """끝쪽 빈 행·열 제거."""
    while rows and all(cell == "" for cell in rows[-1]):
        rows.pop()
    if not rows:
        return rows
    n_cols = max(len(r) for r in rows)
    last_used = -1
    for c in range(n_cols):
        if any(c < len(r) and r[c] != "" for r in rows):
            last_used = c
    if last_used < 0:
        return []
    trimmed = [r[: last_used + 1] for r in rows]
    width = last_used + 1
    return [r + [""] * (width - len(r)) for r in trimmed]


def _rows_to_markdown(rows: list[list[str]]) -> str:
    """2D 행 리스트를 markdown 표로 변환. 1행은 헤더."""
    if not rows:
        return "_(빈 시트)_"
    width = len(rows[0])
    header = rows[0]
    if all(cell == "" for cell in header):
        header = [f"col{i + 1}" for i in range(width)]
        body = rows
    else:
        header = [h if h else f"col{i + 1}" for i, h in enumerate(header)]
        body = rows[1:]
    lines = [
        "| " + " | ".join(header) + " |",
        "| " + " | ".join(["---"] * width) + " |",
    ]
    for row in body:
        padded = row + [""] * (width - len(row))
        lines.append("| " + " | ".join(padded) + " |")
    return "\n".join(lines)


def convert(input_path: Path, sheet: str | None = None) -> str:
    """xlsx 파일 전체(또는 특정 시트)를 markdown 문자열로 반환."""
    wb = load_workbook(input_path, data_only=True, read_only=False)
    if sheet is not None:
        if sheet not in wb.sheetnames:
            raise ValueError(
                f"시트 '{sheet}' 를 찾을 수 없습니다. 사용 가능한 시트: {wb.sheetnames}"
            )
        sheet_names = [sheet]
    else:
        sheet_names = list(wb.sheetnames)

    parts: list[str] = [f"# {input_path.stem}", ""]
    for name in sheet_names:
        ws = wb[name]
        rows = _trim_empty_edges(_sheet_to_rows(ws))
        parts.append(f"## {name}")
        parts.append("")
        parts.append(_rows_to_markdown(rows))
        parts.append("")
    return "\n".join(parts).rstrip() + "\n"


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Excel(.xlsx) 파일을 markdown 으로 변환합니다."
    )
    parser.add_argument("input", type=Path, help="입력 .xlsx 파일")
    parser.add_argument(
        "--out", type=Path, default=None, help="출력 .md 경로 (기본: 입력과 같은 폴더)"
    )
    parser.add_argument("--sheet", default=None, help="특정 시트 이름만 추출")
    args = parser.parse_args()

    if not args.input.exists():
        sys.stderr.write(f"입력 파일을 찾을 수 없습니다: {args.input}\n")
        return 1
    if args.input.suffix.lower() != ".xlsx":
        sys.stderr.write(
            f"확장자가 .xlsx 가 아닙니다: {args.input.suffix}\n"
            "(.xls 는 지원하지 않습니다. Excel 에서 .xlsx 로 저장 후 시도하세요.)\n"
        )
        return 1

    output_path = args.out or args.input.with_suffix(".md")

    try:
        markdown = convert(args.input, sheet=args.sheet)
    except ValueError as e:
        sys.stderr.write(f"{e}\n")
        return 1

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(markdown, encoding="utf-8")
    print(f"변환 완료: {output_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
